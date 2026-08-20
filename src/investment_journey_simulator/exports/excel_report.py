"""Workbook export with summaries, fund tables and series."""

from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

DASHBOARD_SHEET_NAME_STR: str = "Dashboard"
TITLE_FONT_SIZE_INT: int = 16
HEADING_FONT_SIZE_INT: int = 12
NARRATIVE_COLUMN_WIDTH_INT: int = 140
FIRST_COLUMN_WIDTH_INT: int = 28
DATA_COLUMN_WIDTH_INT: int = 20
DATA_COLUMN_LETTERS_STR: str = "BCDEFGHIJKLMNOPQRSTUVWXYZ"
HEADER_FILL_COLOUR_STR: str = "FFE8EAF6"
RUPEE_NUMBER_FORMAT_STR: str = '#,##0'
PERCENT_NUMBER_FORMAT_STR: str = '0.00'
MONEY_KEYWORDS_TUPLE: tuple = (
    "value", "invested", "withdrawn", "tax", "gain", "loss",
    "basis", "charge", "sip", "swp", "cost", "exemption",
    "contributed", "opening", "closing", "requested", "unmet",
)
PERCENT_KEYWORDS_TUPLE: tuple = ("%", "percent", "weight", "return")


def build_excel_report_bytes(
    dashboard_title_str: str,
    nominal_summary_lines_list: list[str],
    real_summary_lines_list: list[str],
    notes_lines_list: list[str],
    sheet_dataframe_dict: dict[str, pd.DataFrame],
) -> bytes:
    """Build the downloadable workbook for one simulation run.

    Brief:
        Writes a narrative dashboard sheet first and then one sheet
        per supplied table so results stay easy to audit.

    Arguments:
        dashboard_title_str (str): Title of the first sheet.
        nominal_summary_lines_list (List[str]): Nominal totals.
        real_summary_lines_list (List[str]): Inflation-adjusted
            totals.
        notes_lines_list (List[str]): Method notes and cautions.
        sheet_dataframe_dict (Dict[str, pd.DataFrame]): Sheet name
            to table mapping written after the dashboard sheet.

    Returns:
        bytes: Complete workbook ready for a download button.

    Warning:
        Charts are not embedded; the series sheets are provided so
        that the reader can plot them natively in the spreadsheet.
    """
    workbook = Workbook()
    dashboard_sheet = workbook.active
    dashboard_sheet.title = DASHBOARD_SHEET_NAME_STR
    _write_dashboard_sheet(
        dashboard_sheet,
        dashboard_title_str,
        nominal_summary_lines_list,
        real_summary_lines_list,
        notes_lines_list,
    )
    for sheet_name_str, table_dataframe in (
        sheet_dataframe_dict.items()
    ):
        _write_dataframe_sheet(
            workbook, sheet_name_str, table_dataframe
        )
    workbook_buffer = BytesIO()
    workbook.save(workbook_buffer)
    return workbook_buffer.getvalue()


def _write_dashboard_sheet(
    sheet: Worksheet,
    dashboard_title_str: str,
    nominal_summary_lines_list: list[str],
    real_summary_lines_list: list[str],
    notes_lines_list: list[str],
) -> None:
    """Write the narrative first sheet of the workbook.

    Brief:
        Prints the title, both summaries and the method notes in
        one wide, wrapped column.

    Arguments:
        sheet (Worksheet): Sheet being written.
        dashboard_title_str (str): Report title.
        nominal_summary_lines_list (List[str]): Nominal totals.
        real_summary_lines_list (List[str]): Real totals.
        notes_lines_list (List[str]): Method notes and cautions.

    Returns:
        None: The sheet is modified in place.

    Warning:
        Existing content on the sheet is overwritten.
    """
    _write_title(sheet, dashboard_title_str)
    next_row_int = _write_text_block(
        sheet,
        "Portfolio Summary (Nominal)",
        nominal_summary_lines_list,
        3,
    )
    next_row_int = _write_text_block(
        sheet,
        "Inflation-adjusted (Real) Summary",
        real_summary_lines_list,
        next_row_int + 2,
    )
    next_row_int = _write_text_block(
        sheet,
        "Notes, cautions and how to use",
        notes_lines_list,
        next_row_int + 2,
    )
    _style_narrative_column(sheet, next_row_int)


def _write_title(sheet: Worksheet, title_str: str) -> None:
    """Write the bold report title into the first cell.

    Brief:
        Anchors the dashboard sheet with a readable heading.

    Arguments:
        sheet (Worksheet): Sheet being written.
        title_str (str): Report title text.

    Returns:
        None: The sheet is modified in place.

    Warning:
        Overwrites whatever already sits in cell A1.
    """
    sheet["A1"] = title_str
    sheet["A1"].font = Font(bold=True, size=TITLE_FONT_SIZE_INT)


def _write_text_block(
    sheet: Worksheet,
    heading_str: str,
    body_lines_list: list[str],
    start_row_int: int,
) -> int:
    """Write a bold heading followed by its body lines.

    Brief:
        Used for each narrative section of the dashboard sheet.

    Arguments:
        sheet (Worksheet): Sheet being written.
        heading_str (str): Bold section heading.
        body_lines_list (List[str]): Lines under the heading.
        start_row_int (int): Row where the heading is written.

    Returns:
        int: First free row after the block.

    Warning:
        Empty strings are written as blank spacer rows.
    """
    sheet[f"A{start_row_int}"] = heading_str
    sheet[f"A{start_row_int}"].font = Font(
        bold=True, size=HEADING_FONT_SIZE_INT
    )
    current_row_int = start_row_int + 1
    for body_line_str in body_lines_list:
        sheet[f"A{current_row_int}"] = body_line_str
        current_row_int += 1
    return current_row_int


def _style_narrative_column(
    sheet: Worksheet,
    last_row_int: int,
) -> None:
    """Widen and wrap the narrative column of the sheet.

    Brief:
        Long caution lines are unreadable in a default column.

    Arguments:
        sheet (Worksheet): Sheet being styled.
        last_row_int (int): Last row that holds narrative text.

    Returns:
        None: The sheet is modified in place.

    Warning:
        Styling stops at the supplied last row.
    """
    sheet.column_dimensions["A"].width = NARRATIVE_COLUMN_WIDTH_INT
    for row_index_int in range(1, last_row_int + 1):
        sheet[f"A{row_index_int}"].alignment = Alignment(
            wrap_text=True, vertical="top"
        )


def _write_dataframe_sheet(
    workbook: Workbook,
    sheet_name_str: str,
    table_dataframe: pd.DataFrame,
) -> None:
    """Write one table into a new worksheet.

    Brief:
        Header cells are bolded and every cell wraps so that long
        labels stay readable.

    Arguments:
        workbook (Workbook): Workbook receiving the sheet.
        sheet_name_str (str): Name of the new worksheet.
        table_dataframe (pd.DataFrame): Table to write.

    Returns:
        None: The workbook is modified in place.

    Warning:
        Sheet names longer than the Excel limit are truncated by
        the underlying library.
    """
    sheet = workbook.create_sheet(title=sheet_name_str[:31])
    for column_index_int, column_name in enumerate(
        table_dataframe.columns, start=1
    ):
        header_cell = sheet.cell(
            row=1, column=column_index_int, value=str(column_name)
        )
        header_cell.font = Font(bold=True)
        header_cell.fill = PatternFill(
            "solid", fgColor=HEADER_FILL_COLOUR_STR
        )
        header_cell.alignment = Alignment(
            vertical="top", wrap_text=True
        )
    _write_body_rows(sheet, table_dataframe)
    _set_column_widths(sheet)
    _apply_sheet_polish(sheet, table_dataframe)


def _apply_sheet_polish(
    sheet: Worksheet,
    table_dataframe: pd.DataFrame,
) -> None:
    """Freeze the header, add a filter and format numbers.

    Brief:
        Turns a raw dump into a sheet an analyst can actually
        work in: headers stay visible, columns can be filtered and
        rupee columns carry thousands separators.

    Arguments:
        sheet (Worksheet): Sheet being polished.
        table_dataframe (pd.DataFrame): Table already written.

    Returns:
        None: The sheet is modified in place.

    Warning:
        Column types are inferred from the header text.
    """
    if table_dataframe.empty:
        return
    sheet.freeze_panes = "A2"
    last_column_letter_str = get_column_letter(
        len(table_dataframe.columns)
    )
    sheet.auto_filter.ref = (
        f"A1:{last_column_letter_str}{len(table_dataframe) + 1}"
    )
    for column_index_int, column_name in enumerate(
        table_dataframe.columns, start=1
    ):
        number_format_str = _resolve_number_format_str(
            str(column_name)
        )
        if number_format_str is None:
            continue
        column_letter_str = get_column_letter(column_index_int)
        for row_index_int in range(2, len(table_dataframe) + 2):
            sheet[
                f"{column_letter_str}{row_index_int}"
            ].number_format = number_format_str


def _resolve_number_format_str(column_name_str: str):
    """Pick a number format from a column's header text.

    Brief:
        Rupee columns get thousands separators, percentage columns
        get two decimals, everything else is left alone.

    Arguments:
        column_name_str (str): Header text of the column.

    Returns:
        Optional[str]: Excel number format, or None.

    Warning:
        Heuristic; a badly named column is simply left unformatted.
    """
    lowered_name_str = column_name_str.lower()
    for keyword_str in PERCENT_KEYWORDS_TUPLE:
        if keyword_str in lowered_name_str:
            return PERCENT_NUMBER_FORMAT_STR
    for keyword_str in MONEY_KEYWORDS_TUPLE:
        if keyword_str in lowered_name_str:
            return RUPEE_NUMBER_FORMAT_STR
    return None


def _write_body_rows(
    sheet: Worksheet,
    table_dataframe: pd.DataFrame,
) -> None:
    """Write the data rows of a table below its header.

    Brief:
        Every cell wraps and aligns to the top so that multi-line
        labels stay readable.

    Arguments:
        sheet (Worksheet): Sheet being written.
        table_dataframe (pd.DataFrame): Table to write.

    Returns:
        None: The sheet is modified in place.

    Warning:
        Values are written as-is, so object columns may land in
        the workbook as text.
    """
    for row_index_int, row_tuple in enumerate(
        table_dataframe.itertuples(index=False), start=2
    ):
        for column_index_int, cell_value in enumerate(
            row_tuple, start=1
        ):
            body_cell = sheet.cell(
                row=row_index_int,
                column=column_index_int,
                value=cell_value,
            )
            body_cell.alignment = Alignment(
                vertical="top", wrap_text=True
            )


def _set_column_widths(sheet: Worksheet) -> None:
    """Widen the label column and every data column.

    Brief:
        Default widths truncate both fund names and rupee values.

    Arguments:
        sheet (Worksheet): Sheet being styled.

    Returns:
        None: The sheet is modified in place.

    Warning:
        Only columns A to Z are widened.
    """
    sheet.column_dimensions["A"].width = FIRST_COLUMN_WIDTH_INT
    for column_letter_str in DATA_COLUMN_LETTERS_STR:
        sheet.column_dimensions[
            column_letter_str
        ].width = DATA_COLUMN_WIDTH_INT
