"""Tests for tables, charts, narrative, exports and the lab."""

from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook

from conftest import build_test_fund, build_test_settings
from investment_journey_simulator.charts import (
    build_dashboard_figure,
    build_donut_trace,
    build_stepped_line_trace,
)
from investment_journey_simulator.constants import (
    DASHBOARD_TITLE_STR,
    REBALANCE_METHOD_FULL_STR,
    REBALANCE_TARGET_SIP_SPLIT_STR,
    SERIES_MONTH_STR,
    SERIES_PORTFOLIO_VALUE_STR,
    STEPUP_MODE_GLOBAL_STR,
    SUMMARY_FUND_NAME_STR,
    SUMMARY_MONEY_COLUMNS_TUPLE,
    WITHDRAWAL_MODE_FIXED_STR,
)
from investment_journey_simulator.dashboard_run import (
    build_figure_title_str,
    build_real_run,
    simulate_nominal_run,
)
from investment_journey_simulator.engine import PortfolioSimulator
from investment_journey_simulator.exports.excel_report import (
    build_excel_report_bytes,
)
from investment_journey_simulator.exports.pdf_report import (
    IS_PDF_TOOLCHAIN_AVAILABLE_BOOL,
    build_pdf_report_bytes,
)
from investment_journey_simulator.formatting import (
    describe_amount_str,
    describe_annual_rate_str,
    describe_months_str,
)
from investment_journey_simulator.models import (
    RebalanceSettings,
    StepUpSettings,
    WithdrawalSettings,
)
from investment_journey_simulator.narrative import (
    build_mode_description_str,
    build_notes_lines_list,
    build_summary_lines_list,
)
from investment_journey_simulator.rebalancing_lab import (
    LabFundSpecification,
    build_default_scenario_list,
    build_headline_dataframe,
    build_scenario_dataframe,
    format_value_with_share_str,
)
from investment_journey_simulator.tables import (
    build_fund_summary_dataframe,
    build_monthly_series_dataframe,
    format_money_columns_dataframe,
)

FULL_FEATURE_SETTINGS = build_test_settings(
    horizon_years_int=8,
    stepup=StepUpSettings(STEPUP_MODE_GLOBAL_STR, 10.0),
    rebalance=RebalanceSettings(
        is_enabled_bool=True,
        interval_months_int=12,
        method_str=REBALANCE_METHOD_FULL_STR,
        target_mode_str=REBALANCE_TARGET_SIP_SPLIT_STR,
    ),
    withdrawal=WithdrawalSettings(
        is_enabled_bool=True,
        start_month_index_int=48,
        mode_str=WITHDRAWAL_MODE_FIXED_STR,
        fixed_amount_float=2000.0,
    ),
)


@pytest.fixture(scope="module")
def sample_result():
    """One completed two-fund run reused by the module.

    REFERENCE: harness only.
    """
    return PortfolioSimulator(
        [
            build_test_fund("Fund-A", 5000.0, 13.0),
            build_test_fund("Fund-B", 5000.0, 9.0),
        ],
        FULL_FEATURE_SETTINGS,
    ).run()


# ------------------------------------------------------------------
# Tables
# ------------------------------------------------------------------
def test_monthly_series_has_one_row_per_month(
    sample_result,
) -> None:
    """The exported series must match the simulated months.

    REFERENCE: G4-SYNTHETIC. Structural invariant.
    """
    series_dataframe = build_monthly_series_dataframe(sample_result)
    assert len(series_dataframe) == 96
    assert SERIES_MONTH_STR in series_dataframe.columns
    assert series_dataframe[SERIES_PORTFOLIO_VALUE_STR].iloc[
        -1
    ] == pytest.approx(sample_result.ending_value_float)


def test_fund_summary_has_one_row_per_fund(sample_result) -> None:
    """The per-fund table must list every fund exactly once.

    REFERENCE: G4-SYNTHETIC. Structural invariant.
    """
    summary_dataframe = build_fund_summary_dataframe(sample_result)
    assert len(summary_dataframe) == 2
    assert set(summary_dataframe[SUMMARY_FUND_NAME_STR]) == {
        "Fund-A",
        "Fund-B",
    }


def test_money_formatting_leaves_the_numeric_table_intact(
    sample_result,
) -> None:
    """Formatting must return a copy, never mutate the source.

    REFERENCE: G4-SYNTHETIC. Purity contract.
    """
    summary_dataframe = build_fund_summary_dataframe(sample_result)
    display_dataframe = format_money_columns_dataframe(
        summary_dataframe, SUMMARY_MONEY_COLUMNS_TUPLE
    )
    assert isinstance(
        display_dataframe["Ending Value"].iloc[0], str
    )
    assert isinstance(
        summary_dataframe["Ending Value"].iloc[0], float
    )


def test_missing_money_column_is_skipped_silently(
    sample_result,
) -> None:
    """An unknown column name must not raise.

    REFERENCE: G4-SYNTHETIC. Defensive branch.
    """
    summary_dataframe = build_fund_summary_dataframe(sample_result)
    display_dataframe = format_money_columns_dataframe(
        summary_dataframe, ["Not A Column"]
    )
    assert len(display_dataframe) == len(summary_dataframe)


# ------------------------------------------------------------------
# Charts
# ------------------------------------------------------------------
def test_dashboard_figure_has_all_panels(sample_result) -> None:
    """The figure must carry three lines, two bars and three pies.

    REFERENCE: G4-SYNTHETIC. Layout contract of the dashboard.
    """
    figure = build_dashboard_figure(
        build_monthly_series_dataframe(sample_result),
        build_fund_summary_dataframe(sample_result),
        "Test",
    )
    assert len(figure.data) == 8
    assert figure.layout.height == 950


def test_donut_falls_back_when_every_gain_is_negative() -> None:
    """A pie cannot draw negative slices, so it must degrade.

    REFERENCE: G4-SYNTHETIC. Known limitation handled explicitly.
    """
    donut_trace = build_donut_trace(["A", "B"], [-100.0, -50.0])
    assert list(donut_trace.labels) == ["No positive gains"]
    assert list(donut_trace.values) == [1.0]


def test_donut_uses_real_values_when_any_gain_is_positive(
) -> None:
    """Positive gains must be plotted, negatives clipped to zero.

    REFERENCE: G4-SYNTHETIC. Clipping branch.
    """
    donut_trace = build_donut_trace(["A", "B"], [100.0, -50.0])
    assert list(donut_trace.values) == [100.0, 0.0]


def test_dashed_line_trace_is_marked_dotted() -> None:
    """The withdrawal line must be visually distinct.

    REFERENCE: G4-SYNTHETIC. Styling branch.
    """
    dashed_trace = build_stepped_line_trace(
        [1, 2], [10.0, 20.0], "Withdrawn", True
    )
    assert dashed_trace.line.dash == "dot"


# ------------------------------------------------------------------
# Narrative
# ------------------------------------------------------------------
def test_mode_description_reports_every_active_switch() -> None:
    """The description must name the switches that were used.

    REFERENCE: G4-SYNTHETIC. Self-documenting output contract.
    """
    description_str = build_mode_description_str(
        FULL_FEATURE_SETTINGS, 6.0
    )
    assert "Rebalancing: ON" in description_str
    assert "SWP: ON" in description_str
    assert "Global step-up" in description_str


def test_mode_description_reports_disabled_switches() -> None:
    """A passive run must say so explicitly.

    REFERENCE: G4-SYNTHETIC. Opposite branch.
    """
    description_str = build_mode_description_str(
        build_test_settings(), 6.0
    )
    assert "Rebalancing: OFF" in description_str
    assert "SWP: OFF" in description_str


def test_summary_lines_cover_the_four_headline_metrics(
    sample_result,
) -> None:
    """Every summary block must carry the same four metrics.

    REFERENCE: G4-SYNTHETIC. Reporting contract.
    """
    summary_lines_list = build_summary_lines_list(
        sample_result, "Nominal"
    )
    assert len(summary_lines_list) == 4
    assert all(
        line_str.startswith("Nominal")
        for line_str in summary_lines_list
    )


def test_notes_contain_cautions_and_usage_guidance() -> None:
    """The caveats must always travel with the numbers.

    REFERENCE: G4-SYNTHETIC. Reporting contract.
    """
    notes_lines_list = build_notes_lines_list()
    joined_str = " ".join(notes_lines_list)
    assert "Important cautions:" in joined_str
    assert "How to use" in joined_str


# ------------------------------------------------------------------
# Run bundling
# ------------------------------------------------------------------
def test_nominal_and_real_runs_bundle_consistently() -> None:
    """Both bundles must describe the same simulation.

    REFERENCE: G4-SYNTHETIC. Bundling contract.
    """
    nominal_run = simulate_nominal_run(
        [build_test_fund(monthly_sip_float=5000.0)],
        build_test_settings(horizon_years_int=10),
    )
    real_run = build_real_run(nominal_run.result, 6.0)
    assert nominal_run.label_str == "Nominal"
    assert real_run.label_str == "Real"
    assert len(real_run.monthly_series_dataframe) == len(
        nominal_run.monthly_series_dataframe
    )
    assert (
        real_run.result.ending_value_float
        < nominal_run.result.ending_value_float
    )


def test_figure_title_lists_the_four_totals(sample_result) -> None:
    """The chart headline must be readable on its own.

    REFERENCE: G4-SYNTHETIC. Export readability contract.
    """
    title_str = build_figure_title_str("Nominal", sample_result)
    for label_str in (
        "End Value",
        "Invested",
        "Withdrawn",
        "Tax Paid",
    ):
        assert label_str in title_str


# ------------------------------------------------------------------
# Exports
# ------------------------------------------------------------------
def test_excel_export_produces_a_readable_workbook(
    sample_result,
) -> None:
    """The workbook must open and carry the expected sheets.

    REFERENCE: G4-SYNTHETIC. Round trip through openpyxl.
    """
    workbook_bytes = build_excel_report_bytes(
        dashboard_title_str=DASHBOARD_TITLE_STR,
        nominal_summary_lines_list=["Nominal End Value: 1"],
        real_summary_lines_list=["Real End Value: 1"],
        notes_lines_list=build_notes_lines_list(),
        sheet_dataframe_dict={
            "Funds": build_fund_summary_dataframe(sample_result),
            "Series": build_monthly_series_dataframe(sample_result),
        },
    )
    workbook = load_workbook(BytesIO(workbook_bytes))
    assert workbook.sheetnames == ["Dashboard", "Funds", "Series"]
    assert workbook["Dashboard"]["A1"].value == DASHBOARD_TITLE_STR


@pytest.mark.skipif(
    not IS_PDF_TOOLCHAIN_AVAILABLE_BOOL,
    reason="reportlab is not installed",
)
def test_pdf_export_produces_a_document(sample_result) -> None:
    """The report must render to real bytes with a header.

    REFERENCE: G4-SYNTHETIC. Round trip through reportlab and
    kaleido; skipped when the optional toolchain is absent.
    """
    figure = build_dashboard_figure(
        build_monthly_series_dataframe(sample_result),
        build_fund_summary_dataframe(sample_result),
        "Test",
    )
    report_bytes = build_pdf_report_bytes(
        dashboard_title_str=DASHBOARD_TITLE_STR,
        nominal_summary_lines_list=["Nominal End Value: 1"],
        real_summary_lines_list=["Real End Value: 1"],
        notes_lines_list=["A note"],
        nominal_figure=figure,
        real_figure=figure,
        nominal_summary_dataframe=build_fund_summary_dataframe(
            sample_result
        ),
        real_summary_dataframe=build_fund_summary_dataframe(
            sample_result
        ),
    )
    assert report_bytes.startswith(b"%PDF")
    assert len(report_bytes) > 10000


# ------------------------------------------------------------------
# Rebalancing laboratory
# ------------------------------------------------------------------
def test_laboratory_compares_nine_policies() -> None:
    """The lab must cover doing nothing plus eight policies.

    REFERENCE: G4-SYNTHETIC. Two methods times two targets times
    two funding choices, plus the passive baseline.
    """
    scenario_list = build_default_scenario_list()
    assert len(scenario_list) == 9
    assert scenario_list[0].is_enabled_bool is False


def test_laboratory_table_reports_funds_and_aggregates() -> None:
    """Each policy table must carry funds and portfolio rows.

    REFERENCE: G4-SYNTHETIC. Reporting contract of the lab.
    """
    fund_specification_list = [
        LabFundSpecification("Fund A", 10000.0, 12.0, 50.0),
        LabFundSpecification("Fund B", 5000.0, 10.0, 50.0),
    ]
    scenario_dataframe = build_scenario_dataframe(
        fund_specification_list,
        build_default_scenario_list()[0],
        10,
        0,
        (1, 5),
    )
    row_label_list = list(scenario_dataframe["Row"])
    assert "Fund A" in row_label_list
    assert "TOTAL value" in row_label_list
    assert "Cumulative tax paid" in row_label_list
    assert list(scenario_dataframe.columns) == ["Row", "T=1Y", "T=5Y"]


def test_laboratory_headline_ranks_every_policy() -> None:
    """The ranking table must carry one row per policy.

    REFERENCE: G4-SYNTHETIC. Reporting contract of the lab.
    """
    headline_dataframe = build_headline_dataframe(
        [LabFundSpecification("Fund A", 10000.0, 12.0, 100.0)],
        10,
        0,
        (5,),
    )
    assert len(headline_dataframe) == 9
    assert "T=5Y value" in headline_dataframe.columns


def test_share_formatter_handles_an_empty_portfolio() -> None:
    """A zero portfolio must report a zero share, not divide.

    REFERENCE: G4-SYNTHETIC. Division-by-zero guard.
    """
    assert "0.00%" in format_value_with_share_str(0.0, 0.0)


# ------------------------------------------------------------------
# Units and magnitudes shown next to every input
# ------------------------------------------------------------------
@pytest.mark.parametrize(
    ("amount_float", "expected_fragment_str"),
    [
        (0.0, "0"),
        (500.0, "500"),
        (5_000.0, "5.00 thousand"),
        (200_000.0, "2.00 lakh"),
        (1_25_000.0, "1.25 lakh"),
        (10_000_000.0, "1.00 crore"),
        (47_59_314.0, "47.59 lakh"),
    ],
)
def test_an_amount_names_its_own_magnitude(
    amount_float: float,
    expected_fragment_str: str,
) -> None:
    """An extra zero must be obvious the moment it is typed.

    REFERENCE: G4-SYNTHETIC. Indian grouping alone still reads as
    a wall of digits, so the magnitude is named as well.

    The thousand tier reads "5.00 thousand" rather than the "5.0"
    this asserted before currencies were configurable. There were
    two spellings of this line - one here and one in currency.py,
    which every other currency already went through - and having
    the dashboard say two different things about the same amount
    was worse than either spelling.
    """
    assert expected_fragment_str in describe_amount_str(
        amount_float
    )


def test_an_amount_description_carries_the_grouped_digits() -> None:
    """The exact figure must survive alongside the friendly name.

    REFERENCE: G4-SYNTHETIC. "2 lakh" alone loses the precision
    the reader typed.
    """
    description_str = describe_amount_str(234567.0)
    assert "2,34,567" in description_str
    assert "2.35 lakh" in description_str


@pytest.mark.parametrize(
    ("months_int", "expected_str"),
    [
        (0, "0 months"),
        (6, "6 months"),
        (12, "12 months - 1.0 years"),
        (240, "240 months - 20.0 years"),
        (600, "600 months - 50.0 years"),
    ],
)
def test_a_duration_is_stated_in_both_units(
    months_int: int,
    expected_str: str,
) -> None:
    """Typed in years, simulated in months - so show both.

    REFERENCE: G4-SYNTHETIC. Under a year, naming years adds
    nothing and is left off.
    """
    assert describe_months_str(months_int) == expected_str


def test_a_rate_names_the_monthly_rate_it_compounds_to() -> None:
    """This is the convention the whole engine rests on.

    REFERENCE: G1-ANALYTIC. 12% a year is 0.9489% a month, not
    1%. Stating it next to the input is how a reader can check
    the tool is doing the right thing.
    """
    description_str = describe_annual_rate_str(12.0)
    assert "12.00% a year" in description_str
    assert "0.9489% a month" in description_str


def test_a_rate_can_be_priced_against_a_principal() -> None:
    """A percentage is abstract until it is money.

    REFERENCE: G4-SYNTHETIC. 12% on 25,000 is 3,000 in the first
    year, which is a figure the reader can sanity-check.
    """
    description_str = describe_annual_rate_str(12.0, 25000.0)
    assert "3,000" in description_str
    assert "25,000" in description_str


def test_a_zero_principal_leaves_the_rate_unpriced() -> None:
    """With nothing to apply it to, no money is invented.

    REFERENCE: G4-SYNTHETIC. Guard branch.
    """
    assert "first year" not in describe_annual_rate_str(12.0, 0.0)


def test_the_monthly_rate_shown_is_the_rate_the_engine_uses(
) -> None:
    """The caption must not drift from the engine's convention.

    REFERENCE: G1-ANALYTIC. Compounding the quoted monthly rate
    twelve times must return the annual rate it came from.
    """
    for annual_percent_float in (6.0, 12.0, 18.0):
        description_str = describe_annual_rate_str(
            annual_percent_float
        )
        monthly_percent_float = float(
            description_str.split("(")[1].split("%")[0]
        )
        compounded_float = (
            (1.0 + monthly_percent_float / 100.0) ** 12 - 1.0
        ) * 100.0
        assert compounded_float == pytest.approx(
            annual_percent_float, abs=0.01
        )
