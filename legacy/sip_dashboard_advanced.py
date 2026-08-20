# sip_dashboard_advanced_full.py
# ------------------------------------------------------------
# Run:
#   pip install -U streamlit plotly pandas openpyxl reportlab kaleido
#   streamlit run sip_dashboard_advanced_full.py
# ------------------------------------------------------------

from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from io import BytesIO

import pandas as pd
import plotly.graph_objects as go
import streamlit as st

# Excel (no images to avoid temp-file image issues)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from plotly.subplots import make_subplots

# PDF snapshot (in-memory images via kaleido)
PDF_OK = True
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Image as RLImage
    from reportlab.platypus import (
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        TableStyle,
    )
    from reportlab.platypus.tables import LongTable
except Exception:
    PDF_OK = False

RUPEE = "₹"

# =========================
# Formatting: ₹ + Indian commas + compact
# =========================
def indian_commas(num: float) -> str:
    n = int(round(float(num)))
    s = str(abs(n))
    if len(s) <= 3:
        out = s
    else:
        last3 = s[-3:]
        rest = s[:-3]
        parts = []
        while len(rest) > 2:
            parts.append(rest[-2:])
            rest = rest[:-2]
        if rest:
            parts.append(rest)
        out = ",".join(reversed(parts)) + "," + last3
    return ("-" if n < 0 else "") + out


def fmt_rupee(num: float) -> str:
    return f"{RUPEE}{indian_commas(num)}"


def inr_compact(num: float) -> str:
    n = float(num)
    sign = "-" if n < 0 else ""
    n = abs(n)
    crore = 10_000_000
    lakh = 100_000
    thousand = 1_000
    if n >= crore:
        return f"{sign}{RUPEE}{n / crore:.2f}Cr"
    if n >= lakh:
        return f"{sign}{RUPEE}{n / lakh:.2f}L"
    if n >= thousand:
        return f"{sign}{RUPEE}{n / thousand:.2f}K"
    return f"{sign}{RUPEE}{indian_commas(n)}"


# =========================
# Dates / rates / helpers
# =========================
def months_between(start: date, end: date) -> int:
    return (end.year - start.year) * 12 + (end.month - start.month)


def monthly_rate_from_annual(annual_percent: float) -> float:
    r = float(annual_percent) / 100.0
    return (1.0 + r) ** (1.0 / 12.0) - 1.0


def net_return_simple(gross_return_percent: float, expense_percent: float) -> float:
    # Planning approximation (net ≈ gross - expense)
    return float(gross_return_percent) - float(expense_percent)


def real_return_from_nominal(nominal_percent: float, inflation_percent: float) -> float:
    nominal = float(nominal_percent) / 100.0
    infl = float(inflation_percent) / 100.0
    if (1.0 + infl) <= 0:
        return float(nominal_percent)
    real = (1.0 + nominal) / (1.0 + infl) - 1.0
    return real * 100.0


def fy_year(d: date) -> int:
    # Indian FY: Apr-Mar. Label by FY start year.
    return d.year if d.month >= 4 else d.year - 1


# =========================
# Lot model for taxation (FIFO)
# =========================
@dataclass
class Lot:
    principal: float
    invest_month: int
    timing_start: bool  # True if invested at start-of-month, else end-of-month


def lot_value(principal: float, i_month: float, end_month: int, invest_month: int, timing_start: bool) -> float:
    held = end_month - invest_month + (1 if timing_start else 0)
    if held <= 0:
        return principal
    return principal * ((1.0 + i_month) ** held)


def lot_basis_sold_for_proceeds(proceeds: float, i_month: float, end_month: int, invest_month: int, timing_start: bool) -> float:
    held = end_month - invest_month + (1 if timing_start else 0)
    if held <= 0:
        return proceeds
    gf = (1.0 + i_month) ** held
    return proceeds / gf


def apply_exemption_to_gain(
    gain: float,
    exemption_amount: float,
    used_exemption: dict[tuple[str, int, str], float],  # (fund, fy, scope) -> used
    fund_name: str,
    fy: int,
    scope: str,  # "LTCG_ONLY" or "TOTAL_GAINS"
) -> float:
    exemption_amount = max(0.0, float(exemption_amount))
    key = (fund_name, fy, scope)
    used = used_exemption.get(key, 0.0)
    remaining = max(0.0, exemption_amount - used)
    applied = min(remaining, max(0.0, gain))
    used_exemption[key] = used + applied
    return max(0.0, gain - applied)


def tax_on_sale(
    *,
    fund_name: str,
    always_stcg: bool,
    months_held: int,
    gain: float,
    stcg_percent: float,
    ltcg_percent: float,
    ltcg_threshold_months: int,
    exemption_amount: float,
    exemption_scope: str,
    sale_date: date,
    used_exemption: dict[tuple[str, int, str], float],
) -> tuple[float, float, float]:
    """
    Returns (tax, stcg_gain_part, ltcg_gain_part) for this realized gain chunk.
    Exemption tracked per fund per FY (planning model).
    """
    if gain <= 0:
        return 0.0, 0.0, 0.0

    fy = fy_year(sale_date)

    if always_stcg:
        taxable_gain = gain
        if exemption_scope == "TOTAL_GAINS":
            taxable_gain = apply_exemption_to_gain(taxable_gain, exemption_amount, used_exemption, fund_name, fy, "TOTAL_GAINS")
        tax = taxable_gain * (float(stcg_percent) / 100.0)
        return tax, gain, 0.0

    if months_held >= int(ltcg_threshold_months):
        taxable_gain = gain
        if exemption_scope == "LTCG_ONLY":
            taxable_gain = apply_exemption_to_gain(taxable_gain, exemption_amount, used_exemption, fund_name, fy, "LTCG_ONLY")
        elif exemption_scope == "TOTAL_GAINS":
            taxable_gain = apply_exemption_to_gain(taxable_gain, exemption_amount, used_exemption, fund_name, fy, "TOTAL_GAINS")
        tax = taxable_gain * (float(ltcg_percent) / 100.0)
        return tax, 0.0, gain

    taxable_gain = gain
    if exemption_scope == "TOTAL_GAINS":
        taxable_gain = apply_exemption_to_gain(taxable_gain, exemption_amount, used_exemption, fund_name, fy, "TOTAL_GAINS")
    tax = taxable_gain * (float(stcg_percent) / 100.0)
    return tax, gain, 0.0


# =========================
# Simulation: Step-up + SWP + Rebalancing + Gaps
# =========================
def simulate_portfolio(
    *,
    funds_cfg: list[dict],
    years_horizon: int,
    portfolio_start: date,
    sip_timing_start_of_month: bool,
    # step-up
    step_mode: str,  # "OFF" | "GLOBAL" | "PER_FUND" | "BOTH"
    global_stepup_percent: float,
    # swp
    swp_enabled: bool,
    swp_start_month: int,
    swp_mode: str,  # "FIXED" | "SCHEDULE_12"
    swp_fixed_amount: float,
    swp_schedule_12: list[float],  # Jan..Dec
    swp_annual_change_percent: float,
    swp_change_12: list[float],  # Jan..Dec annual change % (optional)
    # gaps
    sip_pause_months: list[int],
    swp_pause_months: list[int],
    pause_ranges_df: pd.DataFrame | None,
    # rebalancing
    rebalance_enabled: bool,
    rebalance_every_months: int,
    rebalance_mode: str,          # "Full liquidation (exact target split)" | "Partial (sell overweight only)"
    rebalance_ignore_tax: bool,
    rebalance_target_mode: str,   # "INITIAL_SIP_SPLIT" | "TARGET_ALLOC_COLUMN"
) -> dict:
    total_months = int(years_horizon) * 12
    x_dates = pd.date_range(start=portfolio_start, periods=total_months, freq="MS")
    month_dates = [d.date() for d in x_dates]

    sip_pause_set = set(int(x) for x in sip_pause_months)
    swp_pause_set = set(int(x) for x in swp_pause_months)

    def _safe_date(v) -> date | None:
        if isinstance(v, pd.Timestamp):
            return v.date()
        return v if isinstance(v, date) else None

    def in_pause_range(m: int, kind: str) -> bool:
        if pause_ranges_df is None or len(pause_ranges_df) == 0:
            return False
        d = month_dates[m]
        for _, r in pause_ranges_df.iterrows():
            s = _safe_date(r.get("Start"))
            e = _safe_date(r.get("End"))
            ap = str(r.get("Apply To", "BOTH")).upper().strip()
            if s is None or e is None:
                continue
            if (d.year, d.month) < (s.year, s.month):
                continue
            if (d.year, d.month) > (e.year, e.month):
                continue
            if ap == "BOTH" or ap == kind:
                return True
        return False

    def sip_paused(m: int) -> bool:
        return (month_dates[m].month in sip_pause_set) or in_pause_range(m, "SIP")

    def swp_paused(m: int) -> bool:
        return (month_dates[m].month in swp_pause_set) or in_pause_range(m, "SWP")

    lots: dict[str, list[Lot]] = {f["name"]: [] for f in funds_cfg}

    contributed: dict[str, float] = {f["name"]: 0.0 for f in funds_cfg}
    withdrawn_gross: dict[str, float] = {f["name"]: 0.0 for f in funds_cfg}

    tax_paid: dict[str, float] = {f["name"]: 0.0 for f in funds_cfg}
    stcg_realized: dict[str, float] = {f["name"]: 0.0 for f in funds_cfg}
    ltcg_realized: dict[str, float] = {f["name"]: 0.0 for f in funds_cfg}
    used_exemption: dict[tuple[str, int, str], float] = {}

    portfolio_value_series = []
    portfolio_invested_series = []
    portfolio_withdraw_series = []
    portfolio_tax_series = []
    portfolio_monthly_sip_series = []
    portfolio_monthly_swp_series = []

    # precompute monthly rates
    for f in funds_cfg:
        f["i_month"] = monthly_rate_from_annual(f["net_return_percent"])
        f["target_alloc"] = float(f.get("target_alloc", 0.0))

    # --- IMPORTANT: Only define targets if rebalancing is ON ---
    def set_targets_for_rebalance():
        if not rebalance_enabled:
            # Make sure targets are NEVER used anywhere when OFF
            for f in funds_cfg:
                f["target_alloc"] = 0.0
            return

        if rebalance_target_mode == "INITIAL_SIP_SPLIT":
            sip_sum = sum(max(0.0, float(f.get("sip_monthly", 0.0))) for f in funds_cfg)
            if sip_sum > 0:
                for f in funds_cfg:
                    f["target_alloc"] = 100.0 * max(0.0, float(f.get("sip_monthly", 0.0))) / sip_sum
            else:
                # fallback equal split
                for f in funds_cfg:
                    f["target_alloc"] = 100.0 / len(funds_cfg)

        else:  # TARGET_ALLOC_COLUMN
            s = sum(max(0.0, float(f.get("target_alloc", 0.0))) for f in funds_cfg)
            if s <= 1e-9:
                # fallback to initial SIP split if column is empty
                sip_sum = sum(max(0.0, float(f.get("sip_monthly", 0.0))) for f in funds_cfg)
                if sip_sum > 0:
                    for f in funds_cfg:
                        f["target_alloc"] = 100.0 * max(0.0, float(f.get("sip_monthly", 0.0))) / sip_sum
                else:
                    for f in funds_cfg:
                        f["target_alloc"] = 100.0 / len(funds_cfg)

        # normalize
        s2 = sum(max(0.0, float(f.get("target_alloc", 0.0))) for f in funds_cfg)
        if s2 <= 1e-9:
            for f in funds_cfg:
                f["target_alloc"] = 100.0 / len(funds_cfg)
        else:
            for f in funds_cfg:
                f["target_alloc"] = 100.0 * max(0.0, float(f.get("target_alloc", 0.0))) / s2

    set_targets_for_rebalance()

    def fund_value(fund_name: str, end_month: int, i_month: float) -> float:
        return sum(lot_value(lt.principal, i_month, end_month, lt.invest_month, lt.timing_start) for lt in lots[fund_name])

    def all_values(end_month: int) -> dict[str, float]:
        return {f["name"]: fund_value(f["name"], end_month, f["i_month"]) for f in funds_cfg}

    def sell_from_fund(f: dict, end_month: int, sale_proceeds: float) -> tuple[float, float]:
        if sale_proceeds <= 0:
            return 0.0, 0.0

        name = f["name"]
        i_m = f["i_month"]
        remaining = sale_proceeds
        sold_total = 0.0
        tax_total = 0.0

        new_lots: list[Lot] = []
        for lt in lots[name]:
            if remaining <= 1e-9:
                new_lots.append(lt)
                continue

            lt_val = lot_value(lt.principal, i_m, end_month, lt.invest_month, lt.timing_start)
            if lt_val <= 1e-12:
                continue

            take = min(lt_val, remaining)
            basis_sold = lot_basis_sold_for_proceeds(take, i_m, end_month, lt.invest_month, lt.timing_start)
            gain = max(0.0, take - basis_sold)

            months_held = max(0, end_month - lt.invest_month + (1 if lt.timing_start else 0))

            tax, stcg_part, ltcg_part = tax_on_sale(
                fund_name=name,
                always_stcg=f["always_stcg"],
                months_held=months_held,
                gain=gain,
                stcg_percent=f["stcg_percent"],
                ltcg_percent=f["ltcg_percent"],
                ltcg_threshold_months=f["ltcg_threshold_months"],
                exemption_amount=f["exemption_amount"],
                exemption_scope=f["exemption_scope"],
                sale_date=month_dates[end_month],
                used_exemption=used_exemption,
            )

            tax_total += tax
            stcg_realized[name] += stcg_part
            ltcg_realized[name] += ltcg_part

            sold_total += take
            remaining -= take

            held = end_month - lt.invest_month + (1 if lt.timing_start else 0)
            gf = 1.0 if held <= 0 else (1.0 + i_m) ** held
            remaining_principal = max(0.0, (lt_val - take) / gf)

            if remaining_principal > 1e-9:
                new_lots.append(Lot(principal=remaining_principal, invest_month=lt.invest_month, timing_start=lt.timing_start))

        lots[name] = new_lots
        tax_paid[name] += tax_total
        return sold_total, tax_total

    def swp_amount_for_month(m: int) -> float:
        if (not swp_enabled) or m < swp_start_month:
            return 0.0
        if swp_paused(m):
            return 0.0

        years_since_start = (m - swp_start_month) // 12
        global_factor = (1.0 + float(swp_annual_change_percent) / 100.0) ** years_since_start

        if swp_mode == "FIXED":
            return max(0.0, float(swp_fixed_amount) * global_factor)

        month_of_year = month_dates[m].month  # 1..12
        base = float(swp_schedule_12[month_of_year - 1])

        per_month_rate = float(swp_change_12[month_of_year - 1]) if swp_change_12 and len(swp_change_12) == 12 else 0.0
        per_month_factor = (1.0 + per_month_rate / 100.0) ** years_since_start

        return max(0.0, base * global_factor * per_month_factor)

    def sip_for_fund(f: dict, m: int) -> float:
        if sip_paused(m):
            return 0.0

        offset = max(0, months_between(portfolio_start, f["start_date"]))
        if m < offset:
            return 0.0

        months_since_start = m - offset
        year_index = months_since_start // 12

        base = float(f["sip_monthly"])
        factor = 1.0
        if step_mode in ("GLOBAL", "BOTH"):
            factor *= (1.0 + float(global_stepup_percent) / 100.0) ** year_index
        if step_mode in ("PER_FUND", "BOTH"):
            factor *= (1.0 + float(f["fund_stepup_percent"]) / 100.0) ** year_index

        return max(0.0, base * factor)

    # MAIN loop
    for m in range(total_months):
        total_sip_this_month = 0.0

        # 1) SIP at start-of-month
        if sip_timing_start_of_month:
            for f in funds_cfg:
                c = sip_for_fund(f, m)
                if c > 0:
                    lots[f["name"]].append(Lot(principal=c, invest_month=m, timing_start=True))
                    contributed[f["name"]] += c
                    total_sip_this_month += c

        # 2) Values before SWP (end-of-month)
        vals_before = all_values(m)
        total_before = sum(vals_before.values())

        # 3) SWP: sell proportionally from current values
        wd = swp_amount_for_month(m)
        wd_actual_sold = 0.0
        if wd > 0 and total_before > 1e-9:
            for f in funds_cfg:
                name = f["name"]
                v = vals_before[name]
                share = v / total_before if total_before > 0 else 0.0
                sell_amt = wd * share
                sold, _tax_sold = sell_from_fund(f, m, sell_amt)
                wd_actual_sold += sold
                withdrawn_gross[name] += sold

        # 4) Values after SWP
        vals_after_wd = all_values(m)
        total_after_wd = sum(vals_after_wd.values())

        # 5) Rebalance ONLY if enabled
        if rebalance_enabled and rebalance_every_months > 0:
            if (m + 1) % rebalance_every_months == 0 and total_after_wd > 1e-9:
                # ensure targets are current (in case SIP amounts edited)
                set_targets_for_rebalance()

                if rebalance_mode == "Full liquidation (exact target split)":
                    # Sell everything to cash
                    cash = 0.0
                    tax_reb = 0.0
                    vals_cur = vals_after_wd
                    for f in funds_cfg:
                        name = f["name"]
                        sold, tax_sold = sell_from_fund(f, m, vals_cur[name])
                        cash += sold
                        tax_reb += (0.0 if rebalance_ignore_tax else tax_sold)

                    cash_net = max(0.0, cash - tax_reb)

                    # Buy exactly by target weights (internal transfer)
                    for f in funds_cfg:
                        buy_amt = cash_net * (f["target_alloc"] / 100.0)
                        if buy_amt > 1e-9:
                            lots[f["name"]].append(Lot(principal=buy_amt, invest_month=m, timing_start=False))

                else:
                    # Partial: sell overweight, buy underweight
                    vals_now = vals_after_wd
                    total_now = sum(vals_now.values())
                    desired = {f["name"]: total_now * (f["target_alloc"] / 100.0) for f in funds_cfg}

                    # Sell overweight -> cash
                    cash = 0.0
                    tax_reb = 0.0
                    for f in funds_cfg:
                        name = f["name"]
                        diff = vals_now[name] - desired[name]
                        if diff > 1e-6:
                            sold, tax_sold = sell_from_fund(f, m, diff)
                            cash += sold
                            tax_reb += (0.0 if rebalance_ignore_tax else tax_sold)

                    cash_net = max(0.0, cash - tax_reb)

                    # Buy underweight
                    vals_after_sell = all_values(m)
                    need = {}
                    total_need = 0.0
                    for f in funds_cfg:
                        name = f["name"]
                        diff = desired[name] - vals_after_sell[name]
                        if diff > 1e-6:
                            need[name] = diff
                            total_need += diff

                    if cash_net > 0 and total_need > 0:
                        for f in funds_cfg:
                            name = f["name"]
                            if name not in need:
                                continue
                            buy_amt = cash_net * (need[name] / total_need)
                            if buy_amt > 1e-9:
                                lots[name].append(Lot(principal=buy_amt, invest_month=m, timing_start=False))

        # 6) SIP at end-of-month
        if not sip_timing_start_of_month:
            for f in funds_cfg:
                c = sip_for_fund(f, m)
                if c > 0:
                    lots[f["name"]].append(Lot(principal=c, invest_month=m, timing_start=False))
                    contributed[f["name"]] += c
                    total_sip_this_month += c

        # 7) End-of-month record
        vals_end = all_values(m)
        total_end = sum(vals_end.values())

        total_invested = sum(contributed.values())
        total_withdrawn = sum(withdrawn_gross.values())
        total_tax = sum(tax_paid.values())

        portfolio_value_series.append(total_end)
        portfolio_invested_series.append(total_invested)
        portfolio_withdraw_series.append(total_withdrawn)
        portfolio_tax_series.append(total_tax)
        portfolio_monthly_sip_series.append(total_sip_this_month)
        portfolio_monthly_swp_series.append(wd_actual_sold)

    end_vals = all_values(total_months - 1)
    per_fund = []
    for f in funds_cfg:
        name = f["name"]
        end_val = end_vals[name]
        inv = contributed[name]
        wd = withdrawn_gross[name]
        tax = tax_paid[name]
        wealth_generated = end_val + wd
        gain_total = wealth_generated - inv

        per_fund.append({
            "MF": name,
            "Preset": f["preset"],
            "Start": f["start_date"].isoformat(),
            "Net return % (used)": f["net_return_percent"],
            "Invested (principal)": inv,
            "Withdrawn (gross)": wd,
            "Ending Value": end_val,
            "Wealth Generated (End+Withdraw)": wealth_generated,
            "Gain (Wealth-Invested)": gain_total,
            "Tax Paid (realized)": tax,
            "STCG realized (gross gain)": stcg_realized[name],
            "LTCG realized (gross gain)": ltcg_realized[name],
        })

    return {
        "x_dates": x_dates,
        "portfolio_value_series": portfolio_value_series,
        "portfolio_invested_series": portfolio_invested_series,
        "portfolio_withdraw_series": portfolio_withdraw_series,
        "portfolio_tax_series": portfolio_tax_series,
        "portfolio_monthly_sip_series": portfolio_monthly_sip_series,
        "portfolio_monthly_swp_series": portfolio_monthly_swp_series,
        "per_fund_df": pd.DataFrame(per_fund),
        "end_value": portfolio_value_series[-1] if portfolio_value_series else 0.0,
        "end_invested": portfolio_invested_series[-1] if portfolio_invested_series else 0.0,
        "end_withdrawn": portfolio_withdraw_series[-1] if portfolio_withdraw_series else 0.0,
        "end_tax": portfolio_tax_series[-1] if portfolio_tax_series else 0.0,
    }


# =========================
# Plot styling: dense grid + stepped growth
# =========================
def apply_dense_grid(fig, row=1, col=1):
    fig.update_xaxes(
        showgrid=True, gridwidth=1,
        showline=True, linewidth=1, mirror=True,
        ticks="outside", ticklen=6,
        minor=dict(showgrid=True, gridwidth=0.5),
        row=row, col=col,
    )
    fig.update_yaxes(
        showgrid=True, gridwidth=1,
        showline=True, linewidth=1, mirror=True,
        ticks="outside", ticklen=6,
        minor=dict(showgrid=True, gridwidth=0.5),
        row=row, col=col,
    )
    return fig


def build_dashboard_figure(
    *,
    x_dates,
    invested_series,
    value_series,
    withdraw_series,
    sip_series,
    swp_series,
    names,
    invested_vals,
    gain_vals,
    total_vals,
    title,
):
    fig = make_subplots(
        rows=2, cols=2,
        specs=[[{"type": "xy"}, {"type": "domain"}],
               [{"type": "domain"}, {"type": "domain"}]],
        vertical_spacing=0.12,
        horizontal_spacing=0.08,
        subplot_titles=("Growth (Monthly Steps)", "Invested Split", "Gains Split (positive-only)", "End Value Split")
    )

    fig.add_trace(
        go.Scatter(
            x=x_dates, y=invested_series,
            name="Invested (Principal)",
            mode="lines", line_shape="hv",
            line=dict(width=2.5),
            hovertemplate="%{x|%b %Y}<br>Invested: %{customdata}<extra></extra>",
            customdata=[fmt_rupee(v) for v in invested_series],
        ),
        row=1, col=1
    )
    fig.add_trace(
        go.Scatter(
            x=x_dates, y=value_series,
            name="Portfolio Value (End)",
            mode="lines", line_shape="hv",
            line=dict(width=2.5),
            hovertemplate="%{x|%b %Y}<br>Value: %{customdata}<extra></extra>",
            customdata=[fmt_rupee(v) for v in value_series],
        ),
        row=1, col=1
    )
    fig.add_trace(
        go.Scatter(
            x=x_dates, y=withdraw_series,
            name="Withdrawn (Cumulative, Gross)",
            mode="lines", line_shape="hv",
            line=dict(width=2.5, dash="dot"),
            hovertemplate="%{x|%b %Y}<br>Withdrawn: %{customdata}<extra></extra>",
            customdata=[fmt_rupee(v) for v in withdraw_series],
        ),
        row=1, col=1
    )

    fig.add_trace(
        go.Bar(
            x=x_dates, y=sip_series,
            name="Monthly SIP (Total)",
            opacity=0.25,
            hovertemplate="%{x|%b %Y}<br>Monthly SIP: %{customdata}<extra></extra>",
            customdata=[fmt_rupee(v) for v in sip_series],
        ),
        row=1, col=1
    )

    fig.add_trace(
        go.Bar(
            x=x_dates, y=swp_series,
            name="Monthly SWP (Sold, Gross)",
            opacity=0.25,
            hovertemplate="%{x|%b %Y}<br>Monthly SWP: %{customdata}<extra></extra>",
            customdata=[fmt_rupee(v) for v in swp_series],
        ),
        row=1, col=1
    )

    fig.update_xaxes(title_text="Month", row=1, col=1)
    fig.update_yaxes(title_text=f"Amount ({RUPEE})", row=1, col=1)
    fig.update_layout(hovermode="x unified")
    apply_dense_grid(fig, row=1, col=1)

    def donut(labels, values, hole=0.58):
        vals_pos = [max(0.0, float(v)) for v in values]
        if sum(vals_pos) <= 1e-9:
            labels2 = ["No positive gains"]
            vals2 = [1.0]
            custom = ["All funds are ≤ 0 gain in this view"]
        else:
            labels2 = labels
            vals2 = vals_pos
            custom = [f"{fmt_rupee(v)} ({inr_compact(v)})" for v in values]

        return go.Pie(
            labels=labels2,
            values=vals2,
            hole=hole,
            textinfo="percent",
            hovertemplate="%{label}<br>%{customdata}<extra></extra>",
            customdata=custom,
            sort=False,
            marker=dict(line=dict(width=1)),
            showlegend=False,
        )

    fig.add_trace(donut(names, invested_vals), row=1, col=2)
    fig.add_trace(donut(names, gain_vals), row=2, col=1)
    fig.add_trace(donut(names, total_vals), row=2, col=2)

    fig.update_layout(
        template="plotly_white",
        barmode="overlay",
        height=950,
        title=title,
        title_x=0.02,
        margin=dict(t=95, l=30, r=30, b=30),
        legend=dict(orientation="h", yanchor="bottom", y=1.03, xanchor="left", x=0.02),
        font=dict(size=14),
    )
    return fig


# =========================
# Excel export (NO images)
# =========================
def _write_df_to_sheet(ws, df: pd.DataFrame, start_row=1, start_col=1):
    r0, c0 = start_row, start_col

    for j, col in enumerate(df.columns, start=c0):
        cell = ws.cell(row=r0, column=j, value=str(col))
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for i, row in enumerate(df.itertuples(index=False), start=r0 + 1):
        for j, val in enumerate(row, start=c0):
            cell = ws.cell(row=i, column=j, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    for col_letter in "BCDEFGHIJKLMNOPQRSTUVWXYZ":
        ws.column_dimensions[col_letter].width = 20


def build_excel_bytes(
    *,
    dashboard_title: str,
    summary_lines_nominal: list[str],
    summary_lines_real: list[str],
    notes_lines: list[str],
    funds_table_df: pd.DataFrame,
    nominal_summary_df: pd.DataFrame,
    real_summary_df: pd.DataFrame,
    nominal_series_df: pd.DataFrame,
    real_series_df: pd.DataFrame,
) -> bytes:
    wb = Workbook()
    ws_dash = wb.active
    ws_dash.title = "Dashboard"

    ws_dash["A1"] = dashboard_title
    ws_dash["A1"].font = Font(bold=True, size=16)

    ws_dash["A3"] = "Portfolio Summary (Nominal)"
    ws_dash["A3"].font = Font(bold=True, size=12)
    r = 4
    for line in summary_lines_nominal:
        ws_dash[f"A{r}"] = line
        r += 1

    r += 2
    ws_dash[f"A{r}"] = "Inflation-adjusted (Real) Summary"
    ws_dash[f"A{r}"].font = Font(bold=True, size=12)
    r += 1
    for line in summary_lines_real:
        ws_dash[f"A{r}"] = line
        r += 1

    r += 2
    ws_dash[f"A{r}"] = "Notes / Cautions / How to use"
    ws_dash[f"A{r}"].font = Font(bold=True, size=12)
    r += 1
    for line in notes_lines:
        ws_dash[f"A{r}"] = line
        r += 1

    ws_dash.column_dimensions["A"].width = 140
    for rr in range(1, r + 1):
        ws_dash[f"A{rr}"].alignment = Alignment(wrap_text=True, vertical="top")

    def add_sheet(name: str, df: pd.DataFrame):
        ws = wb.create_sheet(title=name)
        _write_df_to_sheet(ws, df, 1, 1)

    add_sheet("Funds", funds_table_df)
    add_sheet("Nominal_Summary", nominal_summary_df)
    add_sheet("Real_Summary", real_summary_df)
    add_sheet("Nominal_Series", nominal_series_df)
    add_sheet("Real_Series", real_series_df)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# =========================
# PDF snapshot export (in-memory images)
# =========================
def fig_to_png_bytes(fig) -> bytes:
    return fig.to_image(format="png", scale=2)


def build_pdf_bytes(
    *,
    dashboard_title: str,
    summary_lines_nominal: list[str],
    summary_lines_real: list[str],
    notes_lines: list[str],
    fig_nom,
    fig_real,
    nominal_summary_df: pd.DataFrame,
    real_summary_df: pd.DataFrame,
) -> bytes:
    if not PDF_OK:
        raise RuntimeError("reportlab not available")

    styles = getSampleStyleSheet()
    story = []

    bio = BytesIO()
    doc = SimpleDocTemplate(
        bio,
        pagesize=landscape(A4),
        leftMargin=1.2 * cm,
        rightMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
        title=dashboard_title,
    )

    story.append(Paragraph(f"<b>{dashboard_title}</b>", styles["Title"]))
    story.append(Spacer(1, 10))

    story.append(Paragraph("<b>Portfolio Summary (Nominal)</b>", styles["Heading2"]))
    for line in summary_lines_nominal:
        story.append(Paragraph(line, styles["BodyText"]))
    story.append(Spacer(1, 10))
    story.append(RLImage(BytesIO(fig_to_png_bytes(fig_nom)), width=26 * cm, height=16 * cm))
    story.append(Spacer(1, 10))

    story.append(Paragraph("<b>Per-Fund Summary (Nominal)</b>", styles["Heading2"]))
    nom_df = nominal_summary_df.copy()
    nom_data = [list(nom_df.columns)] + nom_df.astype(str).values.tolist()
    nom_table = LongTable(nom_data, repeatRows=1)
    nom_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(nom_table)
    story.append(PageBreak())

    story.append(Paragraph("<b>Inflation-adjusted (Real)</b>", styles["Title"]))
    story.append(Spacer(1, 10))

    story.append(Paragraph("<b>Portfolio Summary (Real)</b>", styles["Heading2"]))
    for line in summary_lines_real:
        story.append(Paragraph(line, styles["BodyText"]))
    story.append(Spacer(1, 10))
    story.append(RLImage(BytesIO(fig_to_png_bytes(fig_real)), width=26 * cm, height=16 * cm))
    story.append(Spacer(1, 10))

    story.append(Paragraph("<b>Per-Fund Summary (Real)</b>", styles["Heading2"]))
    real_df = real_summary_df.copy()
    real_data = [list(real_df.columns)] + real_df.astype(str).values.tolist()
    real_table = LongTable(real_data, repeatRows=1)
    real_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(real_table)
    story.append(PageBreak())

    story.append(Paragraph("<b>Notes, cautions, and how to use effectively</b>", styles["Heading2"]))
    for line in notes_lines:
        story.append(Paragraph(line if line else "&nbsp;", styles["BodyText"]))

    doc.build(story)
    return bio.getvalue()


# =========================
# Streamlit UI
# =========================
st.set_page_config(page_title="SIP Portfolio Dashboard", layout="wide")
st.title("SIP Portfolio Dashboard")

# --- Sidebar controls (not exported) ---
with st.sidebar:
    st.header("Global Settings")
    years = st.slider("Investment Horizon (years)", 1, 60, 10, 1)
    sip_timing = st.radio("SIP Timing", ["Start of Month", "End of Month"], index=0)
    sip_start_of_month = (sip_timing == "Start of Month")
    portfolio_start = st.date_input("Portfolio Start Date", value=date.today())

    st.divider()
    st.subheader("Step-up SIP")
    step_mode = st.radio(
        "Step-up mode",
        ["OFF", "GLOBAL", "PER_FUND", "BOTH"],
        index=0,
        help="GLOBAL applies to all funds. PER_FUND uses each fund's own step-up %. BOTH multiplies both factors.",
    )
    global_stepup_percent = st.number_input(
        "Global step-up % per year",
        min_value=0.0, max_value=100.0, value=10.0, step=0.5,
        disabled=(step_mode not in ["GLOBAL", "BOTH"])
    )

    st.divider()
    st.subheader("Inflation")
    inflation_percent = st.number_input("Inflation % per year", min_value=-5.0, max_value=25.0, value=6.0, step=0.5)

    st.divider()
    st.subheader("Start Dates")
    stagger_enabled = st.toggle("Allow different SIP start dates per fund", value=False)

    st.divider()
    st.subheader("Rebalancing")
    rebalance_enabled = st.toggle("Enable periodic rebalancing", value=False)
    rebalance_target_mode = st.radio(
        "Rebalance target",
        ["INITIAL_SIP_SPLIT", "TARGET_ALLOC_COLUMN"],
        index=0,
        disabled=not rebalance_enabled,
        help="INITIAL_SIP_SPLIT = rebalance to your original SIP proportions. TARGET_ALLOC_COLUMN = use 'Target Allocation %' column."
    )
    rebalance_every_months = st.number_input(
        "Rebalance every N months",
        min_value=1, max_value=120, value=12, step=1,
        disabled=not rebalance_enabled
    )
    rebalance_mode = st.radio(
        "Rebalance method",
        ["Full liquidation (exact target split)", "Partial (sell overweight only)"],
        index=0,
        disabled=not rebalance_enabled
    )
    rebalance_ignore_tax = st.toggle(
        "Ignore taxes during rebalancing (planning)",
        value=False,
        disabled=not rebalance_enabled,
        help="If OFF, taxes reduce cash, but target split is enforced on remaining cash."
    )

    st.divider()
    st.subheader("SWP (Withdrawals)")
    swp_enabled = st.toggle("Enable SWP", value=False)
    swp_start_date = st.date_input(
        "SWP Start Date",
        value=portfolio_start,
        disabled=not swp_enabled
    )
    swp_mode = st.radio("SWP mode", ["FIXED", "SCHEDULE_12"], index=0, disabled=not swp_enabled)
    swp_annual_change = st.number_input("SWP annual change % (+/-)", min_value=-50.0, max_value=50.0, value=0.0, step=0.5, disabled=not swp_enabled)
    swp_fixed_amount = st.number_input(
        f"Fixed withdrawal per month ({RUPEE})",
        min_value=0.0, value=20000.0, step=1000.0,
        disabled=not (swp_enabled and swp_mode == "FIXED")
    )

    swp_schedule_12 = [0.0] * 12
    swp_change_12 = [0.0] * 12
    if swp_enabled and swp_mode == "SCHEDULE_12":
        st.caption("Monthly schedule (Jan..Dec), repeats every year")
        cols = st.columns(3)
        month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        sched = []
        for i in range(12):
            with cols[i % 3]:
                sched.append(st.number_input(f"{month_names[i]}", min_value=0.0, value=0.0, step=1000.0, key=f"swp_m_{i}"))
        swp_schedule_12 = sched

        st.caption("Optional: per-month annual change rates (Jan..Dec), repeats yearly")
        cols2 = st.columns(3)
        for i in range(12):
            with cols2[i % 3]:
                swp_change_12[i] = st.number_input(
                    f"{month_names[i]} Δ% / yr",
                    min_value=-50.0, max_value=50.0,
                    value=0.0, step=0.5,
                    key=f"swp_ch_{i}"
                )

    st.divider()
    st.subheader("Gaps / Pauses (SIP & SWP)")
    sip_pause_months = st.multiselect(
        "Pause SIP in these months (repeats yearly)",
        options=list(range(1, 13)),
        default=[],
        format_func=lambda m: ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"][m-1]
    )
    swp_pause_months = st.multiselect(
        "Pause SWP in these months (repeats yearly)",
        options=list(range(1, 13)),
        default=[],
        format_func=lambda m: ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"][m-1]
    )

    st.caption("Add irregular pause ranges (inclusive of start/end months).")
    pause_table = st.data_editor(
        pd.DataFrame([{"Start": portfolio_start, "End": portfolio_start, "Apply To": "BOTH"}]),
        num_rows="dynamic",
        width="stretch",
        column_config={
            "Start": st.column_config.DateColumn(),
            "End": st.column_config.DateColumn(),
            "Apply To": st.column_config.SelectboxColumn(options=["SIP", "SWP", "BOTH"]),
        },
    )

    st.divider()
    st.subheader("Slab rate (Debt preset)")
    slab_rate_percent = st.number_input("Your slab rate %", min_value=0.0, max_value=60.0, value=30.0, step=0.5)

# Derived SWP start month from date
swp_start_month = max(0, months_between(portfolio_start, swp_start_date))

# Defaults (as you provided)
DEFAULT_EQUITY_STCG = 20.0
DEFAULT_EQUITY_LTCG = 12.5
DEFAULT_EQUITY_LTCG_EXEMPT = 125000.0
DEFAULT_EQUITY_LTCG_MONTHS = 12
DEFAULT_EQUITY_EXEMPT_SCOPE = "LTCG_ONLY"
DEFAULT_DEBT_EXEMPT_SCOPE = "TOTAL_GAINS"

def apply_preset(row: pd.Series) -> pd.Series:
    preset = row.get("Fund Type Preset", "Equity-Oriented (Default)")
    override = bool(row.get("Override Preset?", False))
    if override:
        return row

    if preset == "Equity-Oriented (Default)":
        row["STCG %"] = DEFAULT_EQUITY_STCG
        row["LTCG %"] = DEFAULT_EQUITY_LTCG
        row["LTCG threshold (months)"] = DEFAULT_EQUITY_LTCG_MONTHS
        row["Exemption (₹)"] = DEFAULT_EQUITY_LTCG_EXEMPT
        row["Exemption Scope"] = DEFAULT_EQUITY_EXEMPT_SCOPE
    elif preset == "Debt (post Apr 1, 2023) - slab":
        row["STCG %"] = float(slab_rate_percent)
        row["LTCG %"] = 0.0
        row["LTCG threshold (months)"] = 9999
        row["Exemption (₹)"] = 0.0
        row["Exemption Scope"] = DEFAULT_DEBT_EXEMPT_SCOPE
    return row


# =========================
# Session state: funds + allocation mode
# =========================
if "funds_df" not in st.session_state:
    st.session_state.funds_df = pd.DataFrame([
        {
            "MF Name": "Fund-A",
            "Fund Type Preset": "Equity-Oriented (Default)",
            "Override Preset?": False,
            "SIP / month": 2000,
            "Fund Step-up %": 0.0,
            "Return % (gross)": 12.0,
            "Expense %": 0.20,
            "Fund Start": portfolio_start,
            "Target Allocation %": 50.0,
            "STCG %": DEFAULT_EQUITY_STCG,
            "LTCG %": DEFAULT_EQUITY_LTCG,
            "LTCG threshold (months)": DEFAULT_EQUITY_LTCG_MONTHS,
            "Exemption (₹)": DEFAULT_EQUITY_LTCG_EXEMPT,
            "Exemption Scope": DEFAULT_EQUITY_EXEMPT_SCOPE,
        },
        {
            "MF Name": "Fund-B",
            "Fund Type Preset": "Equity-Oriented (Default)",
            "Override Preset?": False,
            "SIP / month": 2000,
            "Fund Step-up %": 0.0,
            "Return % (gross)": 9.0,
            "Expense %": 0.20,
            "Fund Start": portfolio_start,
            "Target Allocation %": 50.0,
            "STCG %": DEFAULT_EQUITY_STCG,
            "LTCG %": DEFAULT_EQUITY_LTCG,
            "LTCG threshold (months)": DEFAULT_EQUITY_LTCG_MONTHS,
            "Exemption (₹)": DEFAULT_EQUITY_LTCG_EXEMPT,
            "Exemption Scope": DEFAULT_EQUITY_EXEMPT_SCOPE,
        },
    ])

if "alloc_df" not in st.session_state:
    st.session_state.alloc_df = pd.DataFrame([
        {"MF Name": "Fund-A", "Allocation %": 50.0},
        {"MF Name": "Fund-B", "Allocation %": 50.0},
    ])

if "total_sip_monthly" not in st.session_state:
    st.session_state.total_sip_monthly = 4000.0


# =========================
# Input mode
# =========================
st.subheader("Mutual Funds Input")
input_mode = st.radio(
    "Choose input mode",
    options=["Per-fund SIP (manual)", "Total SIP → Allocate by %"],
    index=0,
    horizontal=True
)

if input_mode == "Per-fund SIP (manual)":
    st.caption("Enter SIP per MF directly. Add/remove funds as needed.")

    c_add, c_rm, _ = st.columns([1, 1, 4])
    with c_add:
        if st.button("➕ Add Fund", key="add_fund_manual"):
            st.session_state.funds_df = pd.concat([
                st.session_state.funds_df,
                pd.DataFrame([{
                    "MF Name": f"MF-{len(st.session_state.funds_df) + 1}",
                    "Fund Type Preset": "Equity-Oriented (Default)",
                    "Override Preset?": False,
                    "SIP / month": 1000,
                    "Fund Step-up %": 0.0,
                    "Return % (gross)": 12.0,
                    "Expense %": 0.50,
                    "Fund Start": portfolio_start,
                    "Target Allocation %": 0.0,
                    "STCG %": DEFAULT_EQUITY_STCG,
                    "LTCG %": DEFAULT_EQUITY_LTCG,
                    "LTCG threshold (months)": DEFAULT_EQUITY_LTCG_MONTHS,
                    "Exemption (₹)": DEFAULT_EQUITY_LTCG_EXEMPT,
                    "Exemption Scope": DEFAULT_EQUITY_EXEMPT_SCOPE,
                }])
            ], ignore_index=True)
    with c_rm:
        if st.button("➖ Remove Last Fund", key="rm_fund_manual") and len(st.session_state.funds_df) > 1:
            st.session_state.funds_df = st.session_state.funds_df.iloc[:-1].reset_index(drop=True)

    df_manual = st.session_state.funds_df.copy()
    if not stagger_enabled:
        df_manual["Fund Start"] = portfolio_start
    df_manual = df_manual.apply(apply_preset, axis=1)

    edited_df = st.data_editor(
        df_manual,
        width="stretch",
        num_rows="dynamic",
        column_config={
            "MF Name": st.column_config.TextColumn(required=True),
            "Fund Type Preset": st.column_config.SelectboxColumn(
                options=["Equity-Oriented (Default)", "Debt (post Apr 1, 2023) - slab", "Custom"],
                required=True
            ),
            "Override Preset?": st.column_config.CheckboxColumn(),
            "SIP / month": st.column_config.NumberColumn(min_value=0, step=100),
            "Fund Step-up %": st.column_config.NumberColumn(min_value=0.0, max_value=100.0, step=0.5),
            "Return % (gross)": st.column_config.NumberColumn(step=0.1),
            "Expense %": st.column_config.NumberColumn(min_value=0.0, step=0.05),
            "Fund Start": st.column_config.DateColumn(),
            "Target Allocation %": st.column_config.NumberColumn(min_value=0.0, max_value=100.0, step=0.5),
            "STCG %": st.column_config.NumberColumn(min_value=0.0, max_value=60.0, step=0.5),
            "LTCG %": st.column_config.NumberColumn(min_value=0.0, max_value=60.0, step=0.5),
            "LTCG threshold (months)": st.column_config.NumberColumn(min_value=1, max_value=240, step=1),
            "Exemption (₹)": st.column_config.NumberColumn(min_value=0, step=10000),
            "Exemption Scope": st.column_config.SelectboxColumn(options=["LTCG_ONLY", "TOTAL_GAINS"]),
        }
    )
    st.session_state.funds_df = edited_df

else:
    st.caption("Enter TOTAL SIP per month and split across funds by Allocation %. Start date is the same for all funds.")

    total_sip = st.number_input(
        f"Total SIP per month ({RUPEE})",
        min_value=0.0, step=500.0,
        value=float(st.session_state.total_sip_monthly),
        key="total_sip_input",
    )
    st.session_state.total_sip_monthly = float(total_sip)

    c_add2, c_rm2, c_norm2 = st.columns([1, 1, 2])
    with c_add2:
        if st.button("➕ Add MF row", key="add_alloc_row"):
            st.session_state.alloc_df = pd.concat([
                st.session_state.alloc_df,
                pd.DataFrame([{"MF Name": f"MF-{len(st.session_state.alloc_df) + 1}", "Allocation %": 0.0}])
            ], ignore_index=True)
    with c_rm2:
        if st.button("➖ Remove last row", key="rm_alloc_row") and len(st.session_state.alloc_df) > 1:
            st.session_state.alloc_df = st.session_state.alloc_df.iloc[:-1].reset_index(drop=True)

    alloc_df = st.session_state.alloc_df.copy()
    edited_alloc = st.data_editor(
        alloc_df,
        width="stretch",
        num_rows="dynamic",
        column_config={
            "MF Name": st.column_config.TextColumn(required=True),
            "Allocation %": st.column_config.NumberColumn(min_value=0.0, max_value=100.0, step=0.5),
        },
        key="alloc_editor"
    )
    edited_alloc["MF Name"] = edited_alloc["MF Name"].astype(str).str.strip().replace("", "Unnamed MF")
    edited_alloc["Allocation %"] = pd.to_numeric(edited_alloc["Allocation %"], errors="coerce").fillna(0.0)

    alloc_sum = float(edited_alloc["Allocation %"].sum())
    with c_norm2:
        normalize = st.toggle(
            f"Normalize allocations to 100% (current sum: {alloc_sum:.2f}%)",
            value=(alloc_sum not in [0.0, 100.0]),
            key="alloc_normalize_toggle"
        )

    if alloc_sum > 0 and normalize and abs(alloc_sum - 100.0) > 1e-6:
        edited_alloc["Allocation %"] = edited_alloc["Allocation %"] * (100.0 / alloc_sum)
        alloc_sum = 100.0

    st.session_state.alloc_df = edited_alloc

    derived_rows = []
    for _, r in edited_alloc.iterrows():
        name = str(r["MF Name"]).strip() or "Unnamed MF"
        pct = float(r["Allocation %"])
        sip_amt = float(total_sip) * (pct / 100.0)
        derived_rows.append({
            "MF Name": name,
            "Fund Type Preset": "Equity-Oriented (Default)",
            "Override Preset?": False,
            "SIP / month": round(sip_amt, 2),
            "Fund Step-up %": 0.0,
            "Return % (gross)": 12.0,
            "Expense %": 0.50,
            "Fund Start": portfolio_start,
            "Target Allocation %": pct,
            "STCG %": DEFAULT_EQUITY_STCG,
            "LTCG %": DEFAULT_EQUITY_LTCG,
            "LTCG threshold (months)": DEFAULT_EQUITY_LTCG_MONTHS,
            "Exemption (₹)": DEFAULT_EQUITY_LTCG_EXEMPT,
            "Exemption Scope": DEFAULT_EQUITY_EXEMPT_SCOPE,
        })

    df_alloc_funds = pd.DataFrame(derived_rows).apply(apply_preset, axis=1)
    st.markdown("### Derived funds (edit return/expense/tax/step-up/target allocation)")
    edited_df = st.data_editor(
        df_alloc_funds,
        width="stretch",
        num_rows="dynamic",
        column_config={
            "MF Name": st.column_config.TextColumn(required=True),
            "Fund Type Preset": st.column_config.SelectboxColumn(
                options=["Equity-Oriented (Default)", "Debt (post Apr 1, 2023) - slab", "Custom"],
                required=True
            ),
            "Override Preset?": st.column_config.CheckboxColumn(),
            "SIP / month": st.column_config.NumberColumn(disabled=True),
            "Fund Step-up %": st.column_config.NumberColumn(min_value=0.0, max_value=100.0, step=0.5),
            "Return % (gross)": st.column_config.NumberColumn(step=0.1),
            "Expense %": st.column_config.NumberColumn(min_value=0.0, step=0.05),
            "Fund Start": st.column_config.DateColumn(disabled=True),
            "Target Allocation %": st.column_config.NumberColumn(min_value=0.0, max_value=100.0, step=0.5),
            "STCG %": st.column_config.NumberColumn(min_value=0.0, max_value=60.0, step=0.5),
            "LTCG %": st.column_config.NumberColumn(min_value=0.0, max_value=60.0, step=0.5),
            "LTCG threshold (months)": st.column_config.NumberColumn(min_value=1, max_value=240, step=1),
            "Exemption (₹)": st.column_config.NumberColumn(min_value=0, step=10000),
            "Exemption Scope": st.column_config.SelectboxColumn(options=["LTCG_ONLY", "TOTAL_GAINS"]),
        },
        key="alloc_derived_editor"
    )
    st.session_state.funds_df = edited_df


# =========================
# Build config
# =========================
master_df = st.session_state.funds_df.copy()
master_df = master_df.apply(apply_preset, axis=1)

if not stagger_enabled:
    master_df["Fund Start"] = portfolio_start
if input_mode != "Per-fund SIP (manual)":
    master_df["Fund Start"] = portfolio_start

funds_cfg_nominal = []
for _, row in master_df.iterrows():
    name = str(row.get("MF Name", "")).strip() or "Unnamed MF"
    sip = float(row.get("SIP / month", 0) or 0)
    fund_step = float(row.get("Fund Step-up %", 0) or 0)
    gross = float(row.get("Return % (gross)", 0) or 0)
    exp = float(row.get("Expense %", 0) or 0)
    net = net_return_simple(gross, exp)

    fstart = row.get("Fund Start", portfolio_start)
    if isinstance(fstart, pd.Timestamp):
        fstart = fstart.date()
    if not isinstance(fstart, date):
        fstart = portfolio_start

    preset = row.get("Fund Type Preset", "Equity-Oriented (Default)")
    always_stcg = (preset == "Debt (post Apr 1, 2023) - slab")

    stcg = float(row.get("STCG %", 0) or 0)
    ltcg = float(row.get("LTCG %", 0) or 0)
    ltcg_months = int(row.get("LTCG threshold (months)", 12) or 12)
    exemption = float(row.get("Exemption (₹)", 0) or 0)
    ex_scope = str(row.get("Exemption Scope", "LTCG_ONLY"))

    target_alloc = float(row.get("Target Allocation %", 0) or 0)

    funds_cfg_nominal.append({
        "name": name,
        "sip_monthly": sip,
        "fund_stepup_percent": fund_step,
        "net_return_percent": net,
        "start_date": fstart,
        "preset": preset,
        "always_stcg": always_stcg,
        "stcg_percent": stcg,
        "ltcg_percent": ltcg,
        "ltcg_threshold_months": ltcg_months,
        "exemption_amount": exemption,
        "exemption_scope": ex_scope,
        "target_alloc": target_alloc,
        "gross": gross,
        "expense": exp,
        "net": net,
    })

if len(funds_cfg_nominal) == 0:
    st.stop()

# =========================
# NOMINAL simulation
# =========================
sim_nom = simulate_portfolio(
    funds_cfg=[dict(f) for f in funds_cfg_nominal],
    years_horizon=years,
    portfolio_start=portfolio_start,
    sip_timing_start_of_month=sip_start_of_month,
    step_mode=step_mode,
    global_stepup_percent=global_stepup_percent,
    swp_enabled=swp_enabled,
    swp_start_month=int(swp_start_month),
    swp_mode=swp_mode,
    swp_fixed_amount=float(swp_fixed_amount),
    swp_schedule_12=list(swp_schedule_12),
    swp_annual_change_percent=float(swp_annual_change),
    swp_change_12=list(swp_change_12),
    sip_pause_months=list(sip_pause_months),
    swp_pause_months=list(swp_pause_months),
    pause_ranges_df=pause_table,
    rebalance_enabled=rebalance_enabled,
    rebalance_every_months=int(rebalance_every_months) if rebalance_enabled else 0,
    rebalance_mode=rebalance_mode,
    rebalance_ignore_tax=rebalance_ignore_tax,
    rebalance_target_mode=rebalance_target_mode if rebalance_enabled else "INITIAL_SIP_SPLIT",
)

df_nom = sim_nom["per_fund_df"].copy()
names = df_nom["MF"].tolist()
invested_vals = df_nom["Invested (principal)"].tolist()
gain_vals = df_nom["Gain (Wealth-Invested)"].tolist()
total_vals = df_nom["Ending Value"].tolist()

title_nom = (
    f"Nominal | End Value: {fmt_rupee(sim_nom['end_value'])} | "
    f"Invested: {fmt_rupee(sim_nom['end_invested'])} | "
    f"Withdrawn: {fmt_rupee(sim_nom['end_withdrawn'])} | "
    f"Tax Paid: {fmt_rupee(sim_nom['end_tax'])}"
)

fig_nom = build_dashboard_figure(
    x_dates=sim_nom["x_dates"],
    invested_series=sim_nom["portfolio_invested_series"],
    value_series=sim_nom["portfolio_value_series"],
    withdraw_series=sim_nom["portfolio_withdraw_series"],
    sip_series=sim_nom["portfolio_monthly_sip_series"],
    swp_series=sim_nom["portfolio_monthly_swp_series"],
    names=names,
    invested_vals=invested_vals,
    gain_vals=gain_vals,
    total_vals=total_vals,
    title=title_nom
)

# =========================
# REAL simulation
# =========================
funds_cfg_real = []
for f in funds_cfg_nominal:
    real_net = real_return_from_nominal(f["net_return_percent"], inflation_percent)
    f2 = dict(f)
    f2["net_return_percent"] = real_net
    funds_cfg_real.append(f2)

sim_real = simulate_portfolio(
    funds_cfg=[dict(f) for f in funds_cfg_real],
    years_horizon=years,
    portfolio_start=portfolio_start,
    sip_timing_start_of_month=sip_start_of_month,
    step_mode=step_mode,
    global_stepup_percent=global_stepup_percent,
    swp_enabled=swp_enabled,
    swp_start_month=int(swp_start_month),
    swp_mode=swp_mode,
    swp_fixed_amount=float(swp_fixed_amount),
    swp_schedule_12=list(swp_schedule_12),
    swp_annual_change_percent=float(swp_annual_change),
    swp_change_12=list(swp_change_12),
    sip_pause_months=list(sip_pause_months),
    swp_pause_months=list(swp_pause_months),
    pause_ranges_df=pause_table,
    rebalance_enabled=rebalance_enabled,
    rebalance_every_months=int(rebalance_every_months) if rebalance_enabled else 0,
    rebalance_mode=rebalance_mode,
    rebalance_ignore_tax=rebalance_ignore_tax,
    rebalance_target_mode=rebalance_target_mode if rebalance_enabled else "INITIAL_SIP_SPLIT",
)

df_real = sim_real["per_fund_df"].copy()
names_r = df_real["MF"].tolist()
invested_vals_r = df_real["Invested (principal)"].tolist()
gain_vals_r = df_real["Gain (Wealth-Invested)"].tolist()
total_vals_r = df_real["Ending Value"].tolist()

title_real = (
    f"Real | Inflation: {inflation_percent:.2f}% | End Value: {fmt_rupee(sim_real['end_value'])} | "
    f"Tax Paid: {fmt_rupee(sim_real['end_tax'])}"
)

fig_real = build_dashboard_figure(
    x_dates=sim_real["x_dates"],
    invested_series=sim_real["portfolio_invested_series"],
    value_series=sim_real["portfolio_value_series"],
    withdraw_series=sim_real["portfolio_withdraw_series"],
    sip_series=sim_real["portfolio_monthly_sip_series"],
    swp_series=sim_real["portfolio_monthly_swp_series"],
    names=names_r,
    invested_vals=invested_vals_r,
    gain_vals=gain_vals_r,
    total_vals=total_vals_r,
    title=title_real
)

# =========================
# Summary + display
# =========================
st.subheader("Portfolio Summary")

mode_line = f"Horizon: {years}y • SIP timing: {sip_timing} • Step-up: {step_mode}"
if step_mode in ["GLOBAL", "BOTH"]:
    mode_line += f" (Global {global_stepup_percent:.2f}%/yr)"
if step_mode in ["PER_FUND", "BOTH"]:
    mode_line += " (Per-fund enabled)"
mode_line += f" • Rebalancing: {'ON' if rebalance_enabled else 'OFF'}"
if rebalance_enabled:
    mode_line += f" (Target={rebalance_target_mode}, {rebalance_mode}, every {int(rebalance_every_months)} months, tax={'ignored' if rebalance_ignore_tax else 'applied'})"
mode_line += f" • SWP: {'ON' if swp_enabled else 'OFF'}"
if swp_enabled:
    mode_line += f" ({swp_mode}, start {swp_start_date.isoformat()}, annual {swp_annual_change:.2f}%)"

st.markdown(
    f"""
- **{mode_line}**
- **Nominal End Value:** {fmt_rupee(sim_nom['end_value'])} ({inr_compact(sim_nom['end_value'])})
- **Nominal Invested (Principal):** {fmt_rupee(sim_nom['end_invested'])} ({inr_compact(sim_nom['end_invested'])})
- **Nominal Withdrawn (Gross):** {fmt_rupee(sim_nom['end_withdrawn'])} ({inr_compact(sim_nom['end_withdrawn'])})
- **Nominal Tax Paid (Realized):** {fmt_rupee(sim_nom['end_tax'])} ({inr_compact(sim_nom['end_tax'])})
"""
)

k1, k2, k3, k4 = st.columns(4)
k1.metric("Nominal End Value", fmt_rupee(sim_nom["end_value"]))
k2.metric("Invested (Principal)", fmt_rupee(sim_nom["end_invested"]))
k3.metric("Withdrawn (Gross)", fmt_rupee(sim_nom["end_withdrawn"]))
k4.metric("Tax Paid (Realized)", fmt_rupee(sim_nom["end_tax"]))

st.plotly_chart(fig_nom, width="stretch")

st.subheader("Per-Fund Summary (Nominal)")
df_nom_display = df_nom.copy()
money_cols = [
    "Invested (principal)", "Withdrawn (gross)", "Ending Value", "Wealth Generated (End+Withdraw)",
    "Gain (Wealth-Invested)", "Tax Paid (realized)", "STCG realized (gross gain)", "LTCG realized (gross gain)"
]
for c in money_cols:
    df_nom_display[c] = df_nom_display[c].apply(fmt_rupee)
st.dataframe(df_nom_display, width="stretch", hide_index=True)

st.divider()
st.header("Inflation-adjusted (Real Returns)")
st.markdown(
    f"""
- **Inflation:** {inflation_percent:.2f}% p.a.
- **Real End Value:** {fmt_rupee(sim_real['end_value'])} ({inr_compact(sim_real['end_value'])})
- **Real Tax Paid (Realized):** {fmt_rupee(sim_real['end_tax'])} ({inr_compact(sim_real['end_tax'])})
"""
)

r1, r2, r3, r4 = st.columns(4)
r1.metric("Real End Value", fmt_rupee(sim_real["end_value"]))
r2.metric("Real Invested", fmt_rupee(sim_real["end_invested"]))
r3.metric("Real Withdrawn", fmt_rupee(sim_real["end_withdrawn"]))
r4.metric("Real Tax Paid", fmt_rupee(sim_real["end_tax"]))

st.plotly_chart(fig_real, width="stretch")

st.subheader("Per-Fund Summary (Real)")
df_real_display = df_real.copy()
for c in money_cols:
    df_real_display[c] = df_real_display[c].apply(fmt_rupee)
st.dataframe(df_real_display, width="stretch", hide_index=True)

# =========================
# Notes / cautions + usage
# =========================
notes_text = (
    "Notes: Net return uses net ≈ gross − expense (planning approximation). "
    "Equity preset uses STCG 20%, LTCG 12.5%, 12-month threshold, and ₹1.25L exemption on LTCG only (default). "
    "Debt preset (post Apr 1, 2023) is modeled as always STCG at slab rate. "
    "This is a simulator for planning; not a tax/financial guarantee."
)

with st.expander("Notes, cautions, and how to use this dashboard effectively", expanded=True):
    st.write(notes_text)
    st.markdown(
        """
**Important cautions:**
- **Returns are assumptions.** We use smooth monthly compounding derived from annual inputs; real markets are volatile.
- **Expense ratio handling is an approximation.** We use **net ≈ gross − expense**.
- **SWP & rebalancing taxes are simulated as immediate realized taxes** (lot-wise FIFO). Real timing/payment may differ.
- **Exit load, STT, cess/surcharge, TER changes, dividends/distribution** are not modeled.

**How to use efficiently:**
1. If you want to *see drift* with rebalancing OFF, set different returns for different funds.
2. Use conservative return assumptions.
3. Enter expense ratios from factsheets/TER.
4. Use global/per-fund step-up based on your plan.
5. Rebalancing ON means: “At each rebalance date, take the entire current portfolio value (principal + gains) and re-allocate to the target proportions.”
"""
    )

# =========================
# Exports
# =========================
st.divider()
st.subheader("Export / Save Results")

dashboard_title = "SIP Portfolio Dashboard"

summary_lines_nominal = [
    mode_line,
    f"Nominal End Value: {fmt_rupee(sim_nom['end_value'])} ({inr_compact(sim_nom['end_value'])})",
    f"Nominal Invested (Principal): {fmt_rupee(sim_nom['end_invested'])} ({inr_compact(sim_nom['end_invested'])})",
    f"Nominal Withdrawn (Gross): {fmt_rupee(sim_nom['end_withdrawn'])} ({inr_compact(sim_nom['end_withdrawn'])})",
    f"Nominal Tax Paid (Realized): {fmt_rupee(sim_nom['end_tax'])} ({inr_compact(sim_nom['end_tax'])})",
]
summary_lines_real = [
    f"Inflation: {inflation_percent:.2f}% p.a.",
    f"Real End Value: {fmt_rupee(sim_real['end_value'])} ({inr_compact(sim_real['end_value'])})",
    f"Real Invested: {fmt_rupee(sim_real['end_invested'])} ({inr_compact(sim_real['end_invested'])})",
    f"Real Withdrawn (Gross): {fmt_rupee(sim_real['end_withdrawn'])} ({inr_compact(sim_real['end_withdrawn'])})",
    f"Real Tax Paid (Realized): {fmt_rupee(sim_real['end_tax'])} ({inr_compact(sim_real['end_tax'])})",
]
notes_lines = [
    notes_text,
    "",
    "Important cautions:",
    "• Returns are assumptions; actual markets fluctuate and the path is not smooth.",
    "• Expense ratio is modeled as net ≈ gross − expense (real TER is deducted more continuously).",
    "• SWP & rebalancing taxes are simulated as immediate realized taxes (lot-wise FIFO).",
    "• Exit load, STT, cess/surcharge, TER changes, dividends/distribution are not modeled.",
    "",
    "How to use efficiently:",
    "1) If you want drift with rebalancing OFF, set different returns for different funds.",
    "2) Use conservative return assumptions.",
    "3) Enter expense ratios from fund factsheets/TER.",
    "4) Use global/per-fund step-up based on your plan.",
    "5) Rebalancing ON reallocates total current value (principal+gains) to target proportions on rebalance dates.",
]

df_nom_series = pd.DataFrame({
    "Month": sim_nom["x_dates"],
    "Portfolio_Value": sim_nom["portfolio_value_series"],
    "Portfolio_Invested": sim_nom["portfolio_invested_series"],
    "Portfolio_Withdrawn": sim_nom["portfolio_withdraw_series"],
    "Portfolio_TaxPaid": sim_nom["portfolio_tax_series"],
    "Monthly_SIP": sim_nom["portfolio_monthly_sip_series"],
    "Monthly_SWP_Sold": sim_nom["portfolio_monthly_swp_series"],
})
df_real_series = pd.DataFrame({
    "Month": sim_real["x_dates"],
    "Portfolio_Value": sim_real["portfolio_value_series"],
    "Portfolio_Invested": sim_real["portfolio_invested_series"],
    "Portfolio_Withdrawn": sim_real["portfolio_withdraw_series"],
    "Portfolio_TaxPaid": sim_real["portfolio_tax_series"],
    "Monthly_SIP": sim_real["portfolio_monthly_sip_series"],
    "Monthly_SWP_Sold": sim_real["portfolio_monthly_swp_series"],
})

excel_bytes = build_excel_bytes(
    dashboard_title=dashboard_title,
    summary_lines_nominal=summary_lines_nominal,
    summary_lines_real=summary_lines_real,
    notes_lines=notes_lines,
    funds_table_df=master_df,
    nominal_summary_df=sim_nom["per_fund_df"],
    real_summary_df=sim_real["per_fund_df"],
    nominal_series_df=df_nom_series,
    real_series_df=df_real_series,
)

pdf_bytes = None
pdf_error = None
try:
    if PDF_OK:
        _ = fig_nom.to_image(format="png")  # kaleido test
        pdf_bytes = build_pdf_bytes(
            dashboard_title=dashboard_title,
            summary_lines_nominal=summary_lines_nominal,
            summary_lines_real=summary_lines_real,
            notes_lines=notes_lines,
            fig_nom=fig_nom,
            fig_real=fig_real,
            nominal_summary_df=sim_nom["per_fund_df"],
            real_summary_df=sim_real["per_fund_df"],
        )
except Exception as e:
    pdf_error = str(e)

c1, c2 = st.columns(2)
with c1:
    st.download_button(
        label="⬇️ Download Excel (tables + monthly series)",
        data=excel_bytes,
        file_name="sip_dashboard_export.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
with c2:
    if pdf_bytes is not None:
        st.download_button(
            label="⬇️ Download PDF (snapshot)",
            data=pdf_bytes,
            file_name="sip_dashboard_export.pdf",
            mime="application/pdf",
        )
    else:
        st.warning(
            "PDF export needs: pip install reportlab kaleido (and a working plotly image export).\n\n"
            f"Details: {pdf_error}"
        )
