# sip_dashboard_final.py
# Run:
#   pip install streamlit plotly pandas openpyxl pillow reportlab kaleido
#   streamlit run sip_dashboard_final.py

from __future__ import annotations

from datetime import date
from io import BytesIO

import pandas as pd
import plotly.graph_objects as go
import streamlit as st

# Excel (no images)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from plotly.subplots import make_subplots

# PDF snapshot (in-memory images)
PDF_OK = True
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Image as RLImage,
    )
    from reportlab.platypus import (
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        TableStyle,
    )
    from reportlab.platypus.tables import LongTable
except Exception:
    PDF_OK = False

RUPEE = "₹"


# =========================
# Formatting: ₹ + Indian commas + compact
# =========================
def indian_commas(num: float) -> str:
    n = int(round(float(num)))
    s = str(abs(n))
    if len(s) <= 3:
        out = s
    else:
        last3 = s[-3:]
        rest = s[:-3]
        parts = []
        while len(rest) > 2:
            parts.append(rest[-2:])
            rest = rest[:-2]
        if rest:
            parts.append(rest)
        out = ",".join(reversed(parts)) + "," + last3
    return ("-" if n < 0 else "") + out


def fmt_rupee(num: float) -> str:
    return f"{RUPEE}{indian_commas(num)}"


def inr_compact(num: float) -> str:
    n = float(num)
    sign = "-" if n < 0 else ""
    n = abs(n)
    crore = 10_000_000
    lakh = 100_000
    thousand = 1_000
    if n >= crore:
        return f"{sign}{RUPEE}{n / crore:.2f}Cr"
    if n >= lakh:
        return f"{sign}{RUPEE}{n / lakh:.2f}L"
    if n >= thousand:
        return f"{sign}{RUPEE}{n / thousand:.2f}K"
    return f"{sign}{RUPEE}{indian_commas(n)}"


# =========================
# Rates / math helpers
# =========================
def monthly_rate_from_annual(annual_percent: float) -> float:
    r = float(annual_percent) / 100.0
    return (1.0 + r) ** (1.0 / 12.0) - 1.0


def months_between(start: date, end: date) -> int:
    return (end.year - start.year) * 12 + (end.month - start.month)


def net_return_simple(gross_return_percent: float, expense_percent: float) -> float:
    # Planning approximation: net ≈ gross - expense
    return float(gross_return_percent) - float(expense_percent)


def real_return_from_nominal(nominal_percent: float, inflation_percent: float) -> float:
    nominal = float(nominal_percent) / 100.0
    infl = float(inflation_percent) / 100.0
    if (1.0 + infl) <= 0:
        return float(nominal_percent)
    real = (1.0 + nominal) / (1.0 + infl) - 1.0
    return real * 100.0


# =========================
# SIP simulation (monthly series + per-month contributions for lot tax)
# =========================
def simulate_sip_series(
    sip_monthly: float,
    annual_return_percent: float,
    years_horizon: int,
    sip_timing_start_of_month: bool,
    portfolio_start: date,
    fund_start: date,
    stepup_enabled: bool,
    stepup_percent_per_year: float,
):
    total_months = int(years_horizon) * 12
    i = monthly_rate_from_annual(annual_return_percent)

    offset = max(0, months_between(portfolio_start, fund_start))
    balance = 0.0
    invested = 0.0

    invested_series = []
    fv_series = []
    contrib_series = []

    for m in range(total_months):
        contrib = 0.0
        if m >= offset:
            months_since_start = m - offset
            if stepup_enabled:
                year_index = months_since_start // 12
                contrib = float(sip_monthly) * ((1.0 + stepup_percent_per_year / 100.0) ** year_index)
            else:
                contrib = float(sip_monthly)

        contrib_series.append(contrib)
        invested += contrib

        if sip_timing_start_of_month:
            balance = (balance + contrib) * (1.0 + i)
        else:
            balance = balance * (1.0 + i) + contrib

        invested_series.append(invested)
        fv_series.append(balance)

    return {
        "months": total_months,
        "monthly_rate": i,
        "invested_series": invested_series,
        "fv_series": fv_series,
        "contrib_series": contrib_series,
        "invested_final": invested_series[-1] if invested_series else 0.0,
        "fv_final": fv_series[-1] if fv_series else 0.0,
        "gain_final": (fv_series[-1] - invested_series[-1]) if fv_series else 0.0,
        "offset_months": offset,
    }


# =========================
# Tax (lot-wise at final redemption)
# - Tax applies ONLY on gains, not principal
# =========================
def compute_tax_lotwise(
    contrib_series,
    monthly_rate_i: float,
    sip_timing_start_of_month: bool,
    stcg_percent: float,
    ltcg_percent: float,
    ltcg_threshold_months: int,
    exemption_amount: float,
    exemption_scope: str,  # "LTCG_ONLY" or "TOTAL_GAINS"
    always_stcg: bool,
):
    total_months = len(contrib_series)
    stcg_gain = 0.0
    ltcg_gain = 0.0

    for t, contrib in enumerate(contrib_series):
        if contrib <= 0:
            continue

        # Holding period in months for each installment
        if sip_timing_start_of_month:
            months_held = total_months - t
        else:
            months_held = total_months - t - 1
        months_held = max(0, months_held)

        fv_lot = contrib * ((1.0 + monthly_rate_i) ** months_held)
        gain_lot = fv_lot - contrib  # ✅ taxable base is gain only

        if always_stcg:
            stcg_gain += gain_lot
        else:
            if months_held >= int(ltcg_threshold_months):
                ltcg_gain += gain_lot
            else:
                stcg_gain += gain_lot

    total_gain = stcg_gain + ltcg_gain

    # Apply exemption
    exemption_amount = max(0.0, float(exemption_amount))
    if exemption_scope == "LTCG_ONLY":
        taxable_stcg = max(0.0, stcg_gain)
        taxable_ltcg = max(0.0, ltcg_gain - exemption_amount)
    else:
        taxable_total = max(0.0, total_gain - exemption_amount)
        if total_gain > 0:
            taxable_stcg = taxable_total * (stcg_gain / total_gain)
            taxable_ltcg = taxable_total * (ltcg_gain / total_gain)
        else:
            taxable_stcg = 0.0
            taxable_ltcg = 0.0

    tax = taxable_stcg * (float(stcg_percent) / 100.0) + taxable_ltcg * (float(ltcg_percent) / 100.0)

    return {
        "tax": tax,
        "stcg_gain": stcg_gain,
        "ltcg_gain": ltcg_gain,
        "total_gain": total_gain,
        "taxable_stcg_gain": taxable_stcg,
        "taxable_ltcg_gain": taxable_ltcg,
    }


# =========================
# Plot styling (dense grid + stepped growth + monthly bars)
# =========================
def apply_dense_grid(fig, row=1, col=1):
    fig.update_xaxes(
        showgrid=True, gridwidth=1,
        showline=True, linewidth=1, mirror=True,
        ticks="outside", ticklen=6,
        minor=dict(showgrid=True, gridwidth=0.5),
        row=row, col=col,
    )
    fig.update_yaxes(
        showgrid=True, gridwidth=1,
        showline=True, linewidth=1, mirror=True,
        ticks="outside", ticklen=6,
        minor=dict(showgrid=True, gridwidth=0.5),
        row=row, col=col,
    )
    return fig


def add_growth_panel(fig, x_dates, invested_series, fv_series, contrib_series, row=1, col=1):
    gains_series = [fv - inv for fv, inv in zip(fv_series, invested_series)]

    # Stacked stepped areas: Invested + Gains, plus Total line, plus monthly SIP bars
    fig.add_trace(
        go.Scatter(
            x=x_dates, y=invested_series,
            name="Invested (Principal)",
            mode="lines", line_shape="hv",
            line=dict(width=2.5),
            fill="tozeroy",
            hovertemplate="%{x|%b %Y}<br>Invested: %{customdata}<extra></extra>",
            customdata=[fmt_rupee(v) for v in invested_series],
        ),
        row=row, col=col
    )
    fig.add_trace(
        go.Scatter(
            x=x_dates, y=gains_series,
            name="Gains",
            mode="lines", line_shape="hv",
            line=dict(width=2.5),
            fill="tonexty",
            hovertemplate="%{x|%b %Y}<br>Gains: %{customdata}<extra></extra>",
            customdata=[fmt_rupee(v) for v in gains_series],
        ),
        row=row, col=col
    )
    fig.add_trace(
        go.Scatter(
            x=x_dates, y=fv_series,
            name="Future Value (Total)",
            mode="lines", line_shape="hv",
            line=dict(width=2.5),
            hovertemplate="%{x|%b %Y}<br>Total FV: %{customdata}<extra></extra>",
            customdata=[fmt_rupee(v) for v in fv_series],
        ),
        row=row, col=col
    )
    fig.add_trace(
        go.Bar(
            x=x_dates, y=contrib_series,
            name="Monthly SIP",
            opacity=0.35,
            hovertemplate="%{x|%b %Y}<br>Monthly SIP: %{customdata}<extra></extra>",
            customdata=[fmt_rupee(v) for v in contrib_series],
        ),
        row=row, col=col
    )

    fig.update_xaxes(title_text="Month", row=row, col=col)
    fig.update_yaxes(title_text=f"Amount ({RUPEE})", row=row, col=col)
    fig.update_layout(hovermode="x unified")
    apply_dense_grid(fig, row=row, col=col)
    return fig


def build_dashboard_figure(
    x_dates,
    portfolio_invested,
    portfolio_fv,
    portfolio_contrib,
    names,
    invested_vals,
    gain_vals,
    total_vals,
    title,
):
    fig = make_subplots(
        rows=2, cols=2,
        specs=[[{"type": "xy"}, {"type": "domain"}],
               [{"type": "domain"}, {"type": "domain"}]],
        vertical_spacing=0.12,
        horizontal_spacing=0.08,
        subplot_titles=("Growth (Monthly Steps)", "Invested Split", "Gains Split", "Total FV Split")
    )

    fig = add_growth_panel(
        fig,
        x_dates=x_dates,
        invested_series=portfolio_invested,
        fv_series=portfolio_fv,
        contrib_series=portfolio_contrib,
        row=1, col=1
    )

    def donut(labels, values, hole=0.58):
        return go.Pie(
            labels=labels,
            values=values,
            hole=hole,
            textinfo="percent",
            hovertemplate="%{label}<br>%{customdata}<extra></extra>",
            customdata=[f"{fmt_rupee(v)} ({inr_compact(v)})" for v in values],
            sort=False,
            marker=dict(line=dict(width=1)),
            showlegend=False,
        )

    fig.add_trace(donut(names, invested_vals), row=1, col=2)
    fig.add_trace(donut(names, gain_vals), row=2, col=1)
    fig.add_trace(donut(names, total_vals), row=2, col=2)

    fig.update_layout(
        template="plotly_white",
        barmode="overlay",
        height=950,
        title=title,
        title_x=0.02,
        margin=dict(t=95, l=30, r=30, b=30),
        legend=dict(orientation="h", yanchor="bottom", y=1.03, xanchor="left", x=0.02),
        font=dict(size=14),
    )
    return fig


# =========================
# Excel export (NO images)
# =========================
def _write_df_to_sheet(ws, df: pd.DataFrame, start_row=1, start_col=1):
    r0, c0 = start_row, start_col

    for j, col in enumerate(df.columns, start=c0):
        cell = ws.cell(row=r0, column=j, value=str(col))
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    for i, row in enumerate(df.itertuples(index=False), start=r0 + 1):
        for j, val in enumerate(row, start=c0):
            cell = ws.cell(row=i, column=j, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Basic column width fit (bounded)
    max_width = 45
    for j, col in enumerate(df.columns, start=c0):
        max_len = len(str(col))
        for i in range(r0 + 1, r0 + 1 + len(df)):
            v = ws.cell(row=i, column=j).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        # Only safe for A-Z; good enough here
        letter = chr(64 + (j if j <= 26 else 26))
        ws.column_dimensions[letter].width = min(max_width, max(10, int(max_len * 0.9)))


def build_excel_bytes_rich(
    *,
    dashboard_title: str,
    summary_lines_nominal: list[str],
    summary_lines_real: list[str],
    notes_lines: list[str],
    funds_table_df: pd.DataFrame,
    nominal_summary_df: pd.DataFrame,
    real_summary_df: pd.DataFrame,
    nominal_series_df: pd.DataFrame,
    real_series_df: pd.DataFrame,
) -> bytes:
    wb = Workbook()
    ws_dash = wb.active
    ws_dash.title = "Dashboard"

    ws_dash["A1"] = dashboard_title
    ws_dash["A1"].font = Font(bold=True, size=16)

    ws_dash["A3"] = "Portfolio Summary (Nominal)"
    ws_dash["A3"].font = Font(bold=True, size=12)
    r = 4
    for line in summary_lines_nominal:
        ws_dash[f"A{r}"] = line
        r += 1

    r += 2
    ws_dash[f"A{r}"] = "Inflation-adjusted (Real Returns) Summary"
    ws_dash[f"A{r}"].font = Font(bold=True, size=12)
    r += 1
    for line in summary_lines_real:
        ws_dash[f"A{r}"] = line
        r += 1

    r += 2
    ws_dash[f"A{r}"] = "Notes / Cautions / How to use"
    ws_dash[f"A{r}"].font = Font(bold=True, size=12)
    r += 1
    for line in notes_lines:
        ws_dash[f"A{r}"] = line
        r += 1

    ws_dash.column_dimensions["A"].width = 120
    for rr in range(1, r + 1):
        ws_dash[f"A{rr}"].alignment = Alignment(wrap_text=True, vertical="top")

    def add_sheet(name: str, df: pd.DataFrame):
        ws = wb.create_sheet(title=name)
        _write_df_to_sheet(ws, df, 1, 1)

    add_sheet("Funds", funds_table_df)
    add_sheet("Nominal_Summary", nominal_summary_df)
    add_sheet("Real_Summary", real_summary_df)
    add_sheet("Nominal_Series", nominal_series_df)
    add_sheet("Real_Series", real_series_df)

    # Charts-ready combined series (easy to chart in Excel)
    ws_cd = wb.create_sheet(title="Charts_Data")
    df_cd = nominal_series_df.copy()
    df_cd = df_cd.rename(columns={
        "Portfolio_Invested": "Nominal_Invested",
        "Portfolio_FutureValue": "Nominal_FutureValue",
        "Portfolio_MonthlySIP": "Nominal_MonthlySIP",
    })
    df_cd["Real_Invested"] = real_series_df["Portfolio_Invested"].values
    df_cd["Real_FutureValue"] = real_series_df["Portfolio_FutureValue"].values
    df_cd["Real_MonthlySIP"] = real_series_df["Portfolio_MonthlySIP"].values
    _write_df_to_sheet(ws_cd, df_cd, 1, 1)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


# =========================
# PDF snapshot export (in-memory images)
# =========================
def fig_to_png_bytes(fig) -> bytes:
    # Requires kaleido installed and working
    return fig.to_image(format="png", scale=2)


def build_pdf_snapshot_bytes(
    *,
    dashboard_title: str,
    summary_lines_nominal: list[str],
    summary_lines_real: list[str],
    notes_lines: list[str],
    fig_nom,
    fig_real,
    nominal_summary_df: pd.DataFrame,
    real_summary_df: pd.DataFrame,
) -> bytes:
    if not PDF_OK:
        raise RuntimeError("reportlab not available")

    styles = getSampleStyleSheet()
    story = []

    bio = BytesIO()
    doc = SimpleDocTemplate(
        bio,
        pagesize=landscape(A4),
        leftMargin=1.2 * cm,
        rightMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
        title=dashboard_title,
    )

    story.append(Paragraph(f"<b>{dashboard_title}</b>", styles["Title"]))
    story.append(Spacer(1, 10))

    story.append(Paragraph("<b>Portfolio Summary (Nominal)</b>", styles["Heading2"]))
    for line in summary_lines_nominal:
        story.append(Paragraph(line, styles["BodyText"]))
    story.append(Spacer(1, 10))

    nom_png_bytes = fig_to_png_bytes(fig_nom)
    story.append(RLImage(BytesIO(nom_png_bytes), width=26 * cm, height=16 * cm))
    story.append(Spacer(1, 10))

    story.append(Paragraph("<b>Per-Fund Summary (Nominal + Tax)</b>", styles["Heading2"]))
    nom_df = nominal_summary_df.copy()
    nom_data = [list(nom_df.columns)] + nom_df.astype(str).values.tolist()
    nom_table = LongTable(nom_data, repeatRows=1)
    nom_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(nom_table)
    story.append(PageBreak())

    story.append(Paragraph("<b>Inflation-adjusted (Real Returns)</b>", styles["Title"]))
    story.append(Spacer(1, 10))

    story.append(Paragraph("<b>Portfolio Summary (Real)</b>", styles["Heading2"]))
    for line in summary_lines_real:
        story.append(Paragraph(line, styles["BodyText"]))
    story.append(Spacer(1, 10))

    real_png_bytes = fig_to_png_bytes(fig_real)
    story.append(RLImage(BytesIO(real_png_bytes), width=26 * cm, height=16 * cm))
    story.append(Spacer(1, 10))

    story.append(Paragraph("<b>Per-Fund Summary (Real + Tax)</b>", styles["Heading2"]))
    real_df = real_summary_df.copy()
    real_data = [list(real_df.columns)] + real_df.astype(str).values.tolist()
    real_table = LongTable(real_data, repeatRows=1)
    real_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(real_table)
    story.append(PageBreak())

    story.append(Paragraph("<b>Notes, cautions, and how to use effectively</b>", styles["Heading2"]))
    for line in notes_lines:
        story.append(Paragraph(line if line else "&nbsp;", styles["BodyText"]))

    doc.build(story)
    return bio.getvalue()


# =========================
# Streamlit App
# =========================
st.set_page_config(page_title="SIP Portfolio Dashboard", layout="wide")
st.title("SIP Portfolio Dashboard")

# --- Sidebar (controls; not exported) ---
with st.sidebar:
    st.header("Global Settings")
    years = st.slider("Investment Horizon (years)", 1, 50, 10, 1)
    sip_timing = st.radio("SIP Timing", ["Start of Month", "End of Month"], index=0)
    sip_start_of_month = (sip_timing == "Start of Month")
    portfolio_start = st.date_input("Portfolio Start Date", value=date.today())

    st.divider()
    st.subheader("Step-up SIP")
    stepup_enabled = st.toggle("Enable Step-up SIP", value=False)
    stepup_percent = st.number_input(
        "Step-up % per year",
        min_value=0.0, max_value=100.0, value=10.0, step=0.5,
        disabled=not stepup_enabled
    )

    st.divider()
    st.subheader("Inflation")
    inflation_percent = st.number_input(
        "Inflation % per year",
        min_value=-5.0, max_value=25.0, value=6.0, step=0.5
    )

    st.divider()
    st.subheader("Start Dates")
    stagger_enabled = st.toggle("Allow different SIP start dates per fund", value=False)

    st.divider()
    st.subheader("Slab rate (used by Debt preset)")
    slab_rate_percent = st.number_input("Your slab rate %", min_value=0.0, max_value=60.0, value=30.0, step=0.5)

# --- Defaults per your current rule set (customizable per MF) ---
DEFAULT_EQUITY_STCG = 20.0
DEFAULT_EQUITY_LTCG = 12.5
DEFAULT_EQUITY_LTCG_EXEMPT = 125000.0
DEFAULT_EQUITY_LTCG_MONTHS = 12
DEFAULT_EQUITY_EXEMPT_SCOPE = "LTCG_ONLY"
DEFAULT_DEBT_EXEMPT_SCOPE = "TOTAL_GAINS"


def apply_preset(row: pd.Series) -> pd.Series:
    preset = row.get("Fund Type Preset", "Equity-Oriented (Default)")
    override = bool(row.get("Override Preset?", False))
    if override:
        return row

    if preset == "Equity-Oriented (Default)":
        row["STCG %"] = DEFAULT_EQUITY_STCG
        row["LTCG %"] = DEFAULT_EQUITY_LTCG
        row["LTCG threshold (months)"] = DEFAULT_EQUITY_LTCG_MONTHS
        row["Exemption (₹)"] = DEFAULT_EQUITY_LTCG_EXEMPT
        row["Exemption Scope"] = DEFAULT_EQUITY_EXEMPT_SCOPE
    elif preset == "Debt (post Apr 1, 2023) - slab":
        row["STCG %"] = float(slab_rate_percent)
        row["LTCG %"] = 0.0
        row["LTCG threshold (months)"] = 9999
        row["Exemption (₹)"] = 0.0
        row["Exemption Scope"] = DEFAULT_DEBT_EXEMPT_SCOPE
    return row


# -------------------------
# Session state: manual funds + allocation mode
# -------------------------
if "funds_df" not in st.session_state:
    st.session_state.funds_df = pd.DataFrame([
        {
            "MF Name": "Nifty 50 Index",
            "Fund Type Preset": "Equity-Oriented (Default)",
            "Override Preset?": False,
            "SIP / month": 1500,
            "Return % (gross)": 12.0,
            "Expense %": 0.20,
            "Fund Start": portfolio_start,
            "STCG %": DEFAULT_EQUITY_STCG,
            "LTCG %": DEFAULT_EQUITY_LTCG,
            "LTCG threshold (months)": DEFAULT_EQUITY_LTCG_MONTHS,
            "Exemption (₹)": DEFAULT_EQUITY_LTCG_EXEMPT,
            "Exemption Scope": DEFAULT_EQUITY_EXEMPT_SCOPE,
        },
        {
            "MF Name": "Debt Fund Example",
            "Fund Type Preset": "Debt (post Apr 1, 2023) - slab",
            "Override Preset?": False,
            "SIP / month": 2000,
            "Return % (gross)": 9.0,
            "Expense %": 0.40,
            "Fund Start": portfolio_start,
            "STCG %": float(slab_rate_percent),
            "LTCG %": 0.0,
            "LTCG threshold (months)": 9999,
            "Exemption (₹)": 0.0,
            "Exemption Scope": DEFAULT_DEBT_EXEMPT_SCOPE,
        },
    ])

if "alloc_df" not in st.session_state:
    st.session_state.alloc_df = pd.DataFrame([
        {"MF Name": "MF-1", "Allocation %": 50.0},
        {"MF Name": "MF-2", "Allocation %": 30.0},
        {"MF Name": "MF-3", "Allocation %": 20.0},
    ])

if "total_sip_monthly" not in st.session_state:
    st.session_state.total_sip_monthly = 5000.0


# -------------------------
# Input mode choice (safe; avoids tab side-effect overwrites)
# -------------------------
st.subheader("Mutual Funds Input")
input_mode = st.segmented_control(
    "Choose input mode",
    options=["Per-fund SIP (manual)", "Total SIP → Allocate by %"],
    default="Per-fund SIP (manual)",
)

# -------------------------
# Build/Update master funds_df from selected input mode
# -------------------------
if input_mode == "Per-fund SIP (manual)":
    st.caption("Enter SIP per MF directly. Add/remove funds as needed.")

    c_add, c_rm, c_hint = st.columns([1, 1, 3])

    with c_add:
        if st.button("➕ Add Fund", key="add_fund_manual"):
            st.session_state.funds_df = pd.concat([
                st.session_state.funds_df,
                pd.DataFrame([{
                    "MF Name": f"MF-{len(st.session_state.funds_df) + 1}",
                    "Fund Type Preset": "Equity-Oriented (Default)",
                    "Override Preset?": False,
                    "SIP / month": 1000,
                    "Return % (gross)": 12.0,
                    "Expense %": 0.50,
                    "Fund Start": portfolio_start,
                    "STCG %": DEFAULT_EQUITY_STCG,
                    "LTCG %": DEFAULT_EQUITY_LTCG,
                    "LTCG threshold (months)": DEFAULT_EQUITY_LTCG_MONTHS,
                    "Exemption (₹)": DEFAULT_EQUITY_LTCG_EXEMPT,
                    "Exemption Scope": DEFAULT_EQUITY_EXEMPT_SCOPE,
                }])
            ], ignore_index=True)

    with c_rm:
        if st.button("➖ Remove Last Fund", key="rm_fund_manual") and len(st.session_state.funds_df) > 1:
            st.session_state.funds_df = st.session_state.funds_df.iloc[:-1].reset_index(drop=True)

    with c_hint:
        st.caption("Presets fill defaults; turn Override ON to edit tax fields manually.")

    df_manual = st.session_state.funds_df.copy()

    if not stagger_enabled:
        df_manual["Fund Start"] = portfolio_start

    df_manual = df_manual.apply(apply_preset, axis=1)

    edited_df = st.data_editor(
        df_manual,
        width="stretch",
        num_rows="dynamic",
        column_config={
            "MF Name": st.column_config.TextColumn(required=True),
            "Fund Type Preset": st.column_config.SelectboxColumn(
                options=["Equity-Oriented (Default)", "Debt (post Apr 1, 2023) - slab", "Custom"],
                required=True
            ),
            "Override Preset?": st.column_config.CheckboxColumn(help="Turn ON to edit tax fields manually."),
            "SIP / month": st.column_config.NumberColumn(min_value=0, step=100),
            "Return % (gross)": st.column_config.NumberColumn(step=0.1),
            "Expense %": st.column_config.NumberColumn(min_value=0.0, step=0.05),
            "Fund Start": st.column_config.DateColumn(),
            "STCG %": st.column_config.NumberColumn(min_value=0.0, max_value=60.0, step=0.5),
            "LTCG %": st.column_config.NumberColumn(min_value=0.0, max_value=60.0, step=0.5),
            "LTCG threshold (months)": st.column_config.NumberColumn(min_value=1, max_value=240, step=1),
            "Exemption (₹)": st.column_config.NumberColumn(min_value=0, step=10000),
            "Exemption Scope": st.column_config.SelectboxColumn(options=["LTCG_ONLY", "TOTAL_GAINS"]),
        }
    )

    st.session_state.funds_df = edited_df

else:
    st.caption("Enter TOTAL SIP per month and split across funds by Allocation %. Start date is the same for all funds.")

    total_sip = st.number_input(
        f"Total SIP per month ({RUPEE})",
        min_value=0.0,
        step=500.0,
        value=float(st.session_state.total_sip_monthly),
        key="total_sip_input",
    )
    st.session_state.total_sip_monthly = float(total_sip)

    c_add2, c_rm2, c_norm2 = st.columns([1, 1, 2])

    with c_add2:
        if st.button("➕ Add MF row", key="add_alloc_row"):
            st.session_state.alloc_df = pd.concat([
                st.session_state.alloc_df,
                pd.DataFrame([{"MF Name": f"MF-{len(st.session_state.alloc_df) + 1}", "Allocation %": 0.0}])
            ], ignore_index=True)

    with c_rm2:
        if st.button("➖ Remove last row", key="rm_alloc_row") and len(st.session_state.alloc_df) > 1:
            st.session_state.alloc_df = st.session_state.alloc_df.iloc[:-1].reset_index(drop=True)

    alloc_df = st.session_state.alloc_df.copy()

    edited_alloc = st.data_editor(
        alloc_df,
        width="stretch",
        num_rows="dynamic",
        column_config={
            "MF Name": st.column_config.TextColumn(required=True),
            "Allocation %": st.column_config.NumberColumn(min_value=0.0, max_value=100.0, step=0.5),
        },
        key="alloc_editor"
    )

    edited_alloc["MF Name"] = edited_alloc["MF Name"].astype(str).str.strip().replace("", "Unnamed MF")
    edited_alloc["Allocation %"] = pd.to_numeric(edited_alloc["Allocation %"], errors="coerce").fillna(0.0)

    alloc_sum = float(edited_alloc["Allocation %"].sum())

    with c_norm2:
        normalize = st.toggle(
            f"Normalize allocations to 100% (current sum: {alloc_sum:.2f}%)",
            value=(alloc_sum not in [0.0, 100.0]),
            key="alloc_normalize_toggle"
        )

    if alloc_sum <= 0.0:
        st.warning("Allocation % total is 0. Please enter positive allocation percentages.")
    else:
        if normalize and abs(alloc_sum - 100.0) > 1e-6:
            edited_alloc["Allocation %"] = edited_alloc["Allocation %"] * (100.0 / alloc_sum)
            alloc_sum = 100.0

        if abs(alloc_sum - 100.0) > 1e-6:
            st.warning(f"Allocations do not sum to 100% (current: {alloc_sum:.2f}%). SIP split uses the given % as-is.")
        else:
            st.success("Allocations sum to 100%.")

    st.session_state.alloc_df = edited_alloc

    # Build derived funds_df (start date same for all)
    derived_rows = []
    for _, r in edited_alloc.iterrows():
        name = str(r["MF Name"]).strip() or "Unnamed MF"
        pct = float(r["Allocation %"])
        sip_amt = float(total_sip) * (pct / 100.0)
        derived_rows.append({
            "MF Name": name,
            "Fund Type Preset": "Equity-Oriented (Default)",
            "Override Preset?": False,
            "SIP / month": round(sip_amt, 2),
            "Return % (gross)": 12.0,
            "Expense %": 0.50,
            "Fund Start": portfolio_start,  # forced same date
            "STCG %": DEFAULT_EQUITY_STCG,
            "LTCG %": DEFAULT_EQUITY_LTCG,
            "LTCG threshold (months)": DEFAULT_EQUITY_LTCG_MONTHS,
            "Exemption (₹)": DEFAULT_EQUITY_LTCG_EXEMPT,
            "Exemption Scope": DEFAULT_EQUITY_EXEMPT_SCOPE,
        })

    df_alloc_funds = pd.DataFrame(derived_rows).apply(apply_preset, axis=1)

    st.markdown("### Derived SIP split (you can edit return/expense/tax fields)")
    edited_df = st.data_editor(
        df_alloc_funds,
        width="stretch",
        num_rows="dynamic",
        column_config={
            "MF Name": st.column_config.TextColumn(required=True),
            "Fund Type Preset": st.column_config.SelectboxColumn(
                options=["Equity-Oriented (Default)", "Debt (post Apr 1, 2023) - slab", "Custom"],
                required=True
            ),
            "Override Preset?": st.column_config.CheckboxColumn(help="Turn ON to edit tax fields manually."),
            "SIP / month": st.column_config.NumberColumn(disabled=True),  # computed from allocation
            "Return % (gross)": st.column_config.NumberColumn(step=0.1),
            "Expense %": st.column_config.NumberColumn(min_value=0.0, step=0.05),
            "Fund Start": st.column_config.DateColumn(disabled=True),  # same for all, fixed
            "STCG %": st.column_config.NumberColumn(min_value=0.0, max_value=60.0, step=0.5),
            "LTCG %": st.column_config.NumberColumn(min_value=0.0, max_value=60.0, step=0.5),
            "LTCG threshold (months)": st.column_config.NumberColumn(min_value=1, max_value=240, step=1),
            "Exemption (₹)": st.column_config.NumberColumn(min_value=0, step=10000),
            "Exemption Scope": st.column_config.SelectboxColumn(options=["LTCG_ONLY", "TOTAL_GAINS"]),
        },
        key="alloc_derived_editor"
    )

    st.session_state.funds_df = edited_df

# Master funds table used for all calculations + exports
master_df = st.session_state.funds_df.copy()
if master_df.empty:
    st.stop()

# Ensure presets applied when override = False
master_df = master_df.apply(apply_preset, axis=1)

# Enforce same start date when allocation mode is selected (your requirement)
if input_mode != "Per-fund SIP (manual)":
    master_df["Fund Start"] = portfolio_start

# If stagger disabled, force start date same in manual mode too
if not stagger_enabled:
    master_df["Fund Start"] = portfolio_start

# -------------------------
# Build fund configs from master_df
# -------------------------
funds = []
for _, row in master_df.iterrows():
    name = str(row.get("MF Name", "")).strip() or "Unnamed MF"
    sip = float(row.get("SIP / month", 0) or 0)
    gross = float(row.get("Return % (gross)", 0) or 0)
    exp = float(row.get("Expense %", 0) or 0)

    fstart = row.get("Fund Start", portfolio_start)
    if isinstance(fstart, pd.Timestamp):
        fstart = fstart.date()
    if not isinstance(fstart, date):
        fstart = portfolio_start

    preset = row.get("Fund Type Preset", "Equity-Oriented (Default)")
    always_stcg = (preset == "Debt (post Apr 1, 2023) - slab")

    stcg = float(row.get("STCG %", 0) or 0)
    ltcg = float(row.get("LTCG %", 0) or 0)
    ltcg_months = int(row.get("LTCG threshold (months)", 12) or 12)
    exemption = float(row.get("Exemption (₹)", 0) or 0)
    exemption_scope = str(row.get("Exemption Scope", "LTCG_ONLY"))

    net = net_return_simple(gross, exp)

    funds.append({
        "name": name,
        "sip": sip,
        "gross": gross,
        "expense": exp,
        "net": net,
        "start": fstart,
        "preset": preset,
        "always_stcg": always_stcg,
        "stcg": stcg,
        "ltcg": ltcg,
        "ltcg_months": ltcg_months,
        "exemption": exemption,
        "exemption_scope": exemption_scope,
    })

# =========================
# NOMINAL + TAX
# =========================
total_months = years * 12
x_dates = pd.date_range(start=portfolio_start, periods=total_months, freq="MS")

results_nom = []
portfolio_invested_nom = [0.0] * total_months
portfolio_fv_nom = [0.0] * total_months
portfolio_contrib_nom = [0.0] * total_months

for f in funds:
    sim = simulate_sip_series(
        sip_monthly=f["sip"],
        annual_return_percent=f["net"],
        years_horizon=years,
        sip_timing_start_of_month=sip_start_of_month,
        portfolio_start=portfolio_start,
        fund_start=f["start"],
        stepup_enabled=stepup_enabled,
        stepup_percent_per_year=stepup_percent if stepup_enabled else 0.0,
    )
    tax_info = compute_tax_lotwise(
        contrib_series=sim["contrib_series"],
        monthly_rate_i=sim["monthly_rate"],
        sip_timing_start_of_month=sip_start_of_month,
        stcg_percent=f["stcg"],
        ltcg_percent=f["ltcg"],
        ltcg_threshold_months=f["ltcg_months"],
        exemption_amount=f["exemption"],
        exemption_scope=f["exemption_scope"],
        always_stcg=f["always_stcg"],
    )

    sim.update(f)
    sim.update(tax_info)
    sim["fv_after_tax"] = max(0.0, sim["fv_final"] - sim["tax"])
    sim["gain_after_tax"] = sim["fv_after_tax"] - sim["invested_final"]
    results_nom.append(sim)

    for idx in range(total_months):
        portfolio_invested_nom[idx] += sim["invested_series"][idx]
        portfolio_fv_nom[idx] += sim["fv_series"][idx]
        portfolio_contrib_nom[idx] += sim["contrib_series"][idx]

portfolio_invested_final_nom = portfolio_invested_nom[-1]
portfolio_fv_final_nom = portfolio_fv_nom[-1]
portfolio_tax_nom = sum(r["tax"] for r in results_nom)
portfolio_fv_after_tax_nom = max(0.0, portfolio_fv_final_nom - portfolio_tax_nom)
portfolio_gain_nom = portfolio_fv_final_nom - portfolio_invested_final_nom

names = [r["name"] for r in results_nom]
invested_vals = [r["invested_final"] for r in results_nom]
gain_vals = [r["gain_final"] for r in results_nom]
total_vals = [r["fv_final"] for r in results_nom]

title_nom = (
    f"Nominal | FV: {fmt_rupee(portfolio_fv_final_nom)} | "
    f"Tax: {fmt_rupee(portfolio_tax_nom)} | "
    f"After-tax FV: {fmt_rupee(portfolio_fv_after_tax_nom)}"
)

fig_nom = build_dashboard_figure(
    x_dates=x_dates,
    portfolio_invested=portfolio_invested_nom,
    portfolio_fv=portfolio_fv_nom,
    portfolio_contrib=portfolio_contrib_nom,
    names=names,
    invested_vals=invested_vals,
    gain_vals=gain_vals,
    total_vals=total_vals,
    title=title_nom
)

# =========================
# Portfolio Summary (few-liners)
# =========================
st.subheader("Portfolio Summary")
st.markdown(
    f"""
- **Horizon:** {years} years  •  **SIP timing:** {sip_timing}  •  **Step-up:** {"ON" if stepup_enabled else "OFF"}{" (" + str(stepup_percent) + "%/yr)" if stepup_enabled else ""}
- **Total Invested:** {fmt_rupee(portfolio_invested_final_nom)} ({inr_compact(portfolio_invested_final_nom)})
- **Total Gain (pre-tax):** {fmt_rupee(portfolio_gain_nom)} ({inr_compact(portfolio_gain_nom)})
- **Estimated Tax:** {fmt_rupee(portfolio_tax_nom)} ({inr_compact(portfolio_tax_nom)})
- **Total Future Value (after-tax):** {fmt_rupee(portfolio_fv_after_tax_nom)} ({inr_compact(portfolio_fv_after_tax_nom)})
"""
)

k1, k2, k3, k4 = st.columns(4)
k1.metric("Total Invested", fmt_rupee(portfolio_invested_final_nom))
k2.metric("Total Gain (pre-tax)", fmt_rupee(portfolio_gain_nom))
k3.metric("Estimated Tax", fmt_rupee(portfolio_tax_nom))
k4.metric("After-tax Future Value", fmt_rupee(portfolio_fv_after_tax_nom))

st.plotly_chart(fig_nom, width="stretch")

# =========================
# Nominal per-fund summary table
# =========================
st.subheader("Per-Fund Summary (Nominal + Tax)")
rows_nom = []
for r in results_nom:
    rows_nom.append({
        "MF": r["name"],
        "Preset": r["preset"],
        "SIP / month": fmt_rupee(r["sip"]),
        "Start": r["start"].isoformat(),
        "Gross%": f"{r['gross']:.2f}",
        "Expense%": f"{r['expense']:.2f}",
        "Net% (approx)": f"{r['net']:.2f}",
        "STCG%": f"{r['stcg']:.2f}",
        "LTCG%": f"{r['ltcg']:.2f}",
        "LTCG months": r["ltcg_months"],
        "Exemption": fmt_rupee(r["exemption"]),
        "Exempt scope": r["exemption_scope"],
        "Invested": fmt_rupee(r["invested_final"]),
        "Gain": fmt_rupee(r["gain_final"]),
        "Tax": fmt_rupee(r["tax"]),
        "After-tax FV": fmt_rupee(r["fv_after_tax"]),
    })
df_nom_summary = pd.DataFrame(rows_nom)
st.dataframe(df_nom_summary, width="stretch", hide_index=True)

# =========================
# REAL RETURNS SECTION
# =========================
st.divider()
st.header("Inflation-adjusted (Real Returns)")

results_real = []
portfolio_invested_real = [0.0] * total_months
portfolio_fv_real = [0.0] * total_months
portfolio_contrib_real = [0.0] * total_months

for f in funds:
    real_net = real_return_from_nominal(f["net"], inflation_percent)

    sim = simulate_sip_series(
        sip_monthly=f["sip"],
        annual_return_percent=real_net,
        years_horizon=years,
        sip_timing_start_of_month=sip_start_of_month,
        portfolio_start=portfolio_start,
        fund_start=f["start"],
        stepup_enabled=stepup_enabled,
        stepup_percent_per_year=stepup_percent if stepup_enabled else 0.0,
    )
    tax_info = compute_tax_lotwise(
        contrib_series=sim["contrib_series"],
        monthly_rate_i=sim["monthly_rate"],
        sip_timing_start_of_month=sip_start_of_month,
        stcg_percent=f["stcg"],
        ltcg_percent=f["ltcg"],
        ltcg_threshold_months=f["ltcg_months"],
        exemption_amount=f["exemption"],
        exemption_scope=f["exemption_scope"],
        always_stcg=f["always_stcg"],
    )

    sim.update(f)
    sim["real_net"] = real_net
    sim.update(tax_info)
    sim["fv_after_tax"] = max(0.0, sim["fv_final"] - sim["tax"])
    results_real.append(sim)

    for idx in range(total_months):
        portfolio_invested_real[idx] += sim["invested_series"][idx]
        portfolio_fv_real[idx] += sim["fv_series"][idx]
        portfolio_contrib_real[idx] += sim["contrib_series"][idx]

portfolio_invested_final_real = portfolio_invested_real[-1]
portfolio_fv_final_real = portfolio_fv_real[-1]
portfolio_tax_real = sum(r["tax"] for r in results_real)
portfolio_fv_after_tax_real = max(0.0, portfolio_fv_final_real - portfolio_tax_real)
portfolio_gain_real = portfolio_fv_final_real - portfolio_invested_final_real

names_r = [r["name"] for r in results_real]
invested_vals_r = [r["invested_final"] for r in results_real]
gain_vals_r = [r["gain_final"] for r in results_real]
total_vals_r = [r["fv_final"] for r in results_real]

title_real = (
    f"Real | Inflation: {inflation_percent:.2f}% | "
    f"FV: {fmt_rupee(portfolio_fv_final_real)} | "
    f"Tax: {fmt_rupee(portfolio_tax_real)} | "
    f"After-tax FV: {fmt_rupee(portfolio_fv_after_tax_real)}"
)

fig_real = build_dashboard_figure(
    x_dates=x_dates,
    portfolio_invested=portfolio_invested_real,
    portfolio_fv=portfolio_fv_real,
    portfolio_contrib=portfolio_contrib_real,
    names=names_r,
    invested_vals=invested_vals_r,
    gain_vals=gain_vals_r,
    total_vals=total_vals_r,
    title=title_real
)

st.markdown(
    f"""
- **Inflation:** {inflation_percent:.2f}% p.a.
- **Total Invested:** {fmt_rupee(portfolio_invested_final_real)} ({inr_compact(portfolio_invested_final_real)})
- **Total Gain (pre-tax):** {fmt_rupee(portfolio_gain_real)} ({inr_compact(portfolio_gain_real)})
- **Estimated Tax:** {fmt_rupee(portfolio_tax_real)} ({inr_compact(portfolio_tax_real)})
- **After-tax Future Value:** {fmt_rupee(portfolio_fv_after_tax_real)} ({inr_compact(portfolio_fv_after_tax_real)})
"""
)

a1, a2, a3, a4 = st.columns(4)
a1.metric("Real Invested", fmt_rupee(portfolio_invested_final_real))
a2.metric("Real Gain (pre-tax)", fmt_rupee(portfolio_gain_real))
a3.metric("Estimated Tax", fmt_rupee(portfolio_tax_real))
a4.metric("After-tax Future Value", fmt_rupee(portfolio_fv_after_tax_real))

st.plotly_chart(fig_real, width="stretch")

st.subheader("Per-Fund Summary (Real + Tax)")
rows_real = []
for r in results_real:
    rows_real.append({
        "MF": r["name"],
        "Preset": r["preset"],
        "SIP / month": fmt_rupee(r["sip"]),
        "Start": r["start"].isoformat(),
        "Real net%": f"{r['real_net']:.2f}",
        "Invested": fmt_rupee(r["invested_final"]),
        "Gain": fmt_rupee(r["gain_final"]),
        "Tax": fmt_rupee(r["tax"]),
        "After-tax FV": fmt_rupee(r["fv_after_tax"]),
    })
df_real_summary = pd.DataFrame(rows_real)
st.dataframe(df_real_summary, width="stretch", hide_index=True)

# =========================
# Notes / Cautions (restore + keep to point 5)
# =========================
notes_text = (
    "Notes: Net return uses net ≈ gross − expense (planning approximation). "
    "Equity preset uses STCG 20%, LTCG 12.5%, 12-month threshold, and ₹1.25L exemption on LTCG only. "
    "Debt preset (post Apr 1, 2023) is modeled as always STCG at slab rate. "
    "This is a simulator for planning; not a tax/financial guarantee."
)

with st.expander("Notes, cautions, and how to use this dashboard effectively", expanded=True):
    st.write(notes_text)
    st.markdown(
        """
**Important cautions:**
- **Returns are assumptions.** This model uses smooth monthly compounding derived from your annual input; real markets are volatile.
- **Expense ratio handling is an approximation.** Funds deduct TER more continuously; we use **net ≈ gross − expense** for planning.
- **Tax shown is simplified.** Real taxation can include cess/surcharge, STT, rule updates, and scheme-category nuances.
- **Tax is applied on gains only** (not principal) in this model, and assumed at final redemption.
- **Exit load, TER changes, dividends/distribution, switches, rebalancing, and partial withdrawals** are not modeled here.

**How to use efficiently:**
1. Use conservative return assumptions.
2. Enter expense ratios from fund factsheets/TER.
3. Use Step-up if your SIP increases annually.
4. Use staggered start dates only if you truly start later.
5. Keep tax presets as defaults unless you have a specific category/tax rule you want to simulate.
"""
    )

# =========================
# Export (Excel rich without images + PDF snapshot)
# =========================
st.divider()
st.subheader("Export / Save Results")

dashboard_title = "SIP Portfolio Dashboard"

summary_lines_nominal = [
    f"Horizon: {years} years • SIP timing: {sip_timing} • Step-up: {'ON' if stepup_enabled else 'OFF'}"
    + (f" ({stepup_percent:.2f}%/yr)" if stepup_enabled else ""),
    f"Total Invested: {fmt_rupee(portfolio_invested_final_nom)} ({inr_compact(portfolio_invested_final_nom)})",
    f"Total Gain (pre-tax): {fmt_rupee(portfolio_gain_nom)} ({inr_compact(portfolio_gain_nom)})",
    f"Estimated Tax: {fmt_rupee(portfolio_tax_nom)} ({inr_compact(portfolio_tax_nom)})",
    f"After-tax Future Value: {fmt_rupee(portfolio_fv_after_tax_nom)} ({inr_compact(portfolio_fv_after_tax_nom)})",
]

summary_lines_real = [
    f"Inflation: {inflation_percent:.2f}% p.a.",
    f"Total Invested: {fmt_rupee(portfolio_invested_final_real)} ({inr_compact(portfolio_invested_final_real)})",
    f"Total Gain (pre-tax): {fmt_rupee(portfolio_gain_real)} ({inr_compact(portfolio_gain_real)})",
    f"Estimated Tax: {fmt_rupee(portfolio_tax_real)} ({inr_compact(portfolio_tax_real)})",
    f"After-tax Future Value: {fmt_rupee(portfolio_fv_after_tax_real)} ({inr_compact(portfolio_fv_after_tax_real)})",
]

notes_lines = [
    notes_text,
    "",
    "Important cautions:",
    "• Returns are assumptions; actual markets fluctuate and the path is not smooth.",
    "• Expense ratio is modeled as net ≈ gross − expense (real TER is deducted more continuously).",
    "• Tax is simplified and assumes final redemption; cess/surcharge, STT, rule changes, and category nuances may apply.",
    "• This model taxes gains only (not principal).",
    "• Exit load, TER changes, dividends/distribution, switches, rebalancing, and partial withdrawals are not modeled.",
    "",
    "How to use efficiently:",
    "1) Use conservative return assumptions.",
    "2) Enter expense ratios from fund factsheets/TER.",
    "3) Use Step-up if your SIP increases annually.",
    "4) Use staggered start dates only if you truly start later.",
    "5) Keep tax presets as defaults unless you have a specific category/tax rule you want to simulate.",
]

funds_table_df = master_df.copy()

df_nom_series = pd.DataFrame({
    "Month": x_dates,
    "Portfolio_Invested": portfolio_invested_nom,
    "Portfolio_FutureValue": portfolio_fv_nom,
    "Portfolio_MonthlySIP": portfolio_contrib_nom,
})
df_real_series = pd.DataFrame({
    "Month": x_dates,
    "Portfolio_Invested": portfolio_invested_real,
    "Portfolio_FutureValue": portfolio_fv_real,
    "Portfolio_MonthlySIP": portfolio_contrib_real,
})

excel_bytes = build_excel_bytes_rich(
    dashboard_title=dashboard_title,
    summary_lines_nominal=summary_lines_nominal,
    summary_lines_real=summary_lines_real,
    notes_lines=notes_lines,
    funds_table_df=funds_table_df,
    nominal_summary_df=df_nom_summary,
    real_summary_df=df_real_summary,
    nominal_series_df=df_nom_series,
    real_series_df=df_real_series,
)

pdf_bytes = None
pdf_error = None
try:
    if PDF_OK:
        # Test kaleido viability quickly
        _ = fig_nom.to_image(format="png")
        pdf_bytes = build_pdf_snapshot_bytes(
            dashboard_title=dashboard_title,
            summary_lines_nominal=summary_lines_nominal,
            summary_lines_real=summary_lines_real,
            notes_lines=notes_lines,
            fig_nom=fig_nom,
            fig_real=fig_real,
            nominal_summary_df=df_nom_summary,
            real_summary_df=df_real_summary,
        )
except Exception as e:
    pdf_error = str(e)

c1, c2 = st.columns(2)

with c1:
    st.download_button(
        label="⬇️ Download Excel (rich: text + all tables + all series)",
        data=excel_bytes,
        file_name="sip_dashboard_export.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

with c2:
    if pdf_bytes is not None:
        st.download_button(
            label="⬇️ Download PDF (main page snapshot)",
            data=pdf_bytes,
            file_name="sip_dashboard_export.pdf",
            mime="application/pdf",
        )
    else:
        st.warning(
            "PDF export needs `pip install reportlab kaleido` and a working plotly image export.\n\n"
            f"Details: {pdf_error}"
        )
